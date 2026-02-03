from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework import status  # ✅ ใช้ตัวนี้จัดการ HTTP Status ทั้งหมด
from rest_framework.pagination import PageNumberPagination
from rest_framework.views import APIView
from django.db.models import Sum, Q, Count, F
from django.db import transaction  # ✅ สำหรับระบบ Checkout
from django.contrib.auth import get_user_model, authenticate
from rest_framework.authtoken.models import Token
# ✅ รวม Model ทุกตัวไว้ในบรรทัดเดียว (ป้องกัน Error)
from .models import Product, Order, OrderItem, AdminLog, ProductImage, Review, StockHistory, Category, Tag, Coupon, FlashSale, FlashSaleProduct, UserCoupon, FlashSaleCampaign
from .serializers import CouponSerializer, FlashSaleSerializer, FlashSaleProductSerializer, FlashSaleCampaignSerializer 
from .validators import validate_order_data
from .exceptions import InlineValidationError 
from .services import CouponService # ✅ Import Service 
import logging
import traceback
from django.utils import timezone
import csv
from django.http import HttpResponse
from datetime import timedelta, datetime
import calendar

User = get_user_model()
logger = logging.getLogger(__name__)

# ==========================================
# 🍌 Mega Menu API
# ==========================================
@api_view(['GET'])
@permission_classes([AllowAny])
def get_menu_configs_api(request):
    """
    API สำหรับดึง Config ของ Mega Menu ทั้งหมด
    Frontend: BananaMenu.jsx
    """
    try:
        categories = Category.objects.all().prefetch_related('menu_config')
        # We need a custom serializer for the list of categories with their config
        from .serializers import CategorySerializer
        serializer = CategorySerializer(categories, many=True)
        return Response(serializer.data)
    except Exception as e:
        logger.error(f"Error fetching menu configs: {e}")
        return Response({'error': str(e)}, status=500)

# ==========================================
# ✅ PromptPay Helper Functions (Native Implementation to avoid Lib Error)
import binascii

def crc16(data: str) -> str:
    crc = 0xFFFF
    for char in data:
        crc ^= ord(char) << 8
        for _ in range(8):
            if (crc & 0x8000):
                crc = (crc << 1) ^ 0x1021
            else:
                crc <<= 1
        crc &= 0xFFFF
    return hex(crc)[2:].upper().zfill(4)

def generate_promptpay_payload(amount):
    """
    Generate PromptPay QR payload (EMVCo) for a specific amount.
    Target ID: 098-765-4321 (Mock ID) -> 0066987654321
    """
    # 1. Payload Format Indicator
    payload = "000201"
    # 2. Point of Initiation Method (12 = Dynamic/Amount included)
    payload += "010212"
    
    # 3. Merchant Account Information (ID 29 for PromptPay)
    # AID: A000000063000101 (Tag 00, Len 16)
    # BillerID/Phone: 0066987654321 (Tag 01, Len 13)
    merchant_data = "0016A000000063000101" + "01130066987654321" 
    payload += "29" + str(len(merchant_data)).zfill(2) + merchant_data
    
    # 4. Currency Code (53) -> 764 (THB)
    payload += "5303764"
    
    # 5. Transaction Amount (54)
    amt_str = "{:.2f}".format(float(amount))
    payload += "54" + str(len(amt_str)).zfill(2) + amt_str
    
    # 6. Country Code (58) -> TH
    payload += "5802TH"
    
    # 7. CRC (63)
    payload += "6304" # Checksum placeholder
    
    # Calculate CRC
    checksum = crc16(payload)
    return payload + checksum
# ==========================================
# 🔧 Admin & Super Admin Core Logic
# ==========================================

class DashboardStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.role not in ['seller', 'admin', 'super_admin']: 
            return Response(status=403)
        
        # 1. Parse Date Filters
        view_mode = request.query_params.get('view_mode', 'daily')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        
        from django.utils import timezone
        import datetime
        from django.db.models import Count, Sum
        from django.db.models.functions import TruncDate, TruncMonth, TruncYear

        # Default range: Last 30 days if not specified
        today = timezone.now().date()
        
        if start_date_str:
            start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
        else:
            start_date = today - datetime.timedelta(days=30)
            
        if end_date_str:
            end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
        else:
            end_date = today

        # 2. Main Queryset (Filter by Date Range)
        # Include Paid, Processing, Shipped, Completed as "Valid Sales"
        valid_statuses = ['Paid', 'Processing', 'Shipped', 'Completed']
        
        orders_in_range = Order.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        )
        
        # 3. Calculate Stats
        sales_agg = orders_in_range.filter(status__in=valid_statuses).aggregate(Sum('total_price'))
        total_sales = sales_agg['total_price__sum'] or 0
        
        print(f"DEBUG: Start={start_date}, End={end_date}")
        print(f"DEBUG: Valid Statuses={valid_statuses}")
        print(f"DEBUG: Orders in Range={orders_in_range.count()}")
        print(f"DEBUG: Paid in Range={orders_in_range.filter(status__in=valid_statuses).count()}")
        print(f"DEBUG: Total Sales Calculated={total_sales}")

        total_orders = orders_in_range.count()
        pending_orders = Order.objects.filter(status='Pending').count() # Pending is usually global/current
        
        # Global Counts (All time)
        global_total_sales = Order.objects.filter(status__in=valid_statuses).aggregate(Sum('total_price'))['total_price__sum'] or 0
        global_total_orders = Order.objects.count()

        # 4. Graph Data
        # Group by Date/Month depending on view_mode
        sales_data = []
        
        if view_mode == 'daily':
            # Generate all dates in range to avoid gaps
            delta = end_date - start_date
            sales_dict = {}
            
            # Query grouped by date
            daily_stats = orders_in_range.filter(status__in=valid_statuses)\
                .annotate(date=TruncDate('created_at'))\
                .values('date')\
                .annotate(sales=Sum('total_price'))\
                .order_by('date')
                
            for stat in daily_stats:
                sales_dict[stat['date']] = stat['sales']
            
            for i in range(delta.days + 1):
                d = start_date + datetime.timedelta(days=i)
                sales_data.append({
                    "name": d.strftime("%d/%m"),
                    "sales": sales_dict.get(d, 0)
                })

        elif view_mode == 'monthly':
            # Logic for monthly grouping (Simplify for now: just daily within the month)
             # ... (reuse daily logic for simplicity or implement TruncMonth)
            delta = end_date - start_date
            sales_dict = {}
            daily_stats = orders_in_range.filter(status__in=valid_statuses)\
                .annotate(date=TruncDate('created_at'))\
                .values('date')\
                .annotate(sales=Sum('total_price'))\
                .order_by('date')
            for stat in daily_stats:
                sales_dict[stat['date']] = stat['sales']
            for i in range(delta.days + 1):
                d = start_date + datetime.timedelta(days=i)
                sales_data.append({
                    "name": d.strftime("%d/%m"),
                    "sales": sales_dict.get(d, 0)
                })
        else: # Yearly (Show sales per month)
             monthly_stats = orders_in_range.filter(status__in=valid_statuses)\
                .annotate(month=TruncMonth('created_at'))\
                .values('month')\
                .annotate(sales=Sum('total_price'))\
                .order_by('month')
             
             sales_map = {m['month'].strftime("%b"): m['sales'] for m in monthly_stats}
             months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
             for m in months:
                 sales_data.append({"name": m, "sales": sales_map.get(m, 0)})

        # 2. Low Stock Products (Stock < 10)
        low_stock_products = Product.objects.filter(stock__lt=10, is_active=True).order_by('stock')[:5]
        low_stock_data = [{
            "id": p.id, 
            "title": p.title, 
            "stock": p.stock, 
            "thumbnail": p.thumbnail.url if p.thumbnail else ""
        } for p in low_stock_products]

        # 5. Pie Chart Data (Sales by Category)
        category_stats = orders_in_range.filter(status__in=valid_statuses)\
            .values('items__product__category')\
            .annotate(value=Sum('items__price_at_purchase'))\
            .order_by('-value')
        
        pie_data = []
        for c in category_stats:
            cat_name = c['items__product__category'] or "Uncategorized"
            pie_data.append({"name": cat_name, "value": c['value']})

        # 6. Bar Chart & Best Sellers (Top 5 Products by Quantity)
        product_stats = orders_in_range.filter(status__in=valid_statuses)\
            .values('items__product__id', 'items__product__title', 'items__product__price', 'items__product__thumbnail')\
            .annotate(sales=Sum('items__quantity'))\
            .order_by('-sales')[:5]
            
        bar_data = []
        for p in product_stats:
            p_name = p['items__product__title'] or "Unknown Product"
            # Encode Thumbnail URL safely
            thumb_url = ""
            if p['items__product__thumbnail']:
                 thumb_url = "/media/" + p['items__product__thumbnail'] # Explicit path construction for values query
            
            bar_data.append({
                "id": p['items__product__id'],
                "name": p_name, # Full name for list
                "short_name": p_name[:15], # Truncated for bar chart
                "price": p['items__product__price'] or 0,
                "sales": p['sales'],
                "thumbnail": thumb_url
            })

        # 7. Activity Logs (Recent 5)
        recent_logs = AdminLog.objects.all().order_by('-timestamp')[:5]
        logs_data = [{
            "id": l.id,
            "action": l.action,
            "date": l.timestamp.strftime("%d/%m %H:%M"),
            "admin": l.admin.username
        } for l in recent_logs]

        # 8. Province Sales Data (Map Analytics) ✅
        province_stats = orders_in_range.filter(status__in=valid_statuses)\
            .values('shipping_province')\
            .annotate(total_sales=Sum('total_price'))\
            .order_by('-total_sales')

        province_data = []
        for p in province_stats:
            prov_name = p['shipping_province'] or "Unknown"
            
            # Find Top Product for this province
            top_prod = orders_in_range.filter(shipping_province=prov_name, status__in=valid_statuses)\
                .values('items__product__title')\
                .annotate(qty=Sum('items__quantity'))\
                .order_by('-qty').first()
                
            top_product_name = top_prod['items__product__title'] if top_prod else "-"
            top_product_qty = top_prod['qty'] if top_prod else 0

            # ✅ ดึงรายชื่อลูกค้าล่าสุด 5 คนในจังหวัดนี้
            recent_buyers_qs = orders_in_range.filter(shipping_province=prov_name, status__in=valid_statuses)\
                .values_list('user__username', flat=True).distinct()[:5]
            
            recent_buyers_list = list(recent_buyers_qs)

            province_data.append({
                "name": prov_name,
                "value": p['total_sales'],
                "top_product": f"{top_product_name} ({top_product_qty})",
                "recent_buyers": recent_buyers_list # ✅ เพิ่มรายชื่อคนซื้อ
            })

        return Response({
            "total_sales": global_total_sales,
            "total_orders": global_total_orders,
            "total_products": Product.objects.count(),
            "total_users": User.objects.count(),
            "pending_orders": pending_orders,
            "low_stock": low_stock_data,
            "sales_data": sales_data,
            "best_sellers": bar_data,
            "pie_data": pie_data,
            "bar_data": bar_data,
            "province_data": province_data, # ✅ Add Province Data
            "logs": logs_data
        })


@api_view(['GET'])
@permission_classes([AllowAny])
def products_api(request):
    try:
        # ดึงสินค้าที่สถานะ is_active=True (เปิดขาย) เท่านั้น
        products = Product.objects.filter(is_active=True)
        
        # ✅ Sorting Logic (การเรียงลำดับ)
        sort_option = request.query_params.get('sort')
        if sort_option == 'price_asc':
            products = products.order_by('price')
        elif sort_option == 'price_desc':
            products = products.order_by('-price')
        else:
            products = products.order_by('-id') # Default (Newest)

        category = request.query_params.get('category')
        search = request.query_params.get('search')
        
        if category and category != "ทั้งหมด":
            products = products.filter(category__name=category)
        
        # ✅ Brand Filter
        brand = request.query_params.get('brand')
        if brand and brand != "ทั้งหมด":
            products = products.filter(brand=brand)
        
        # ✅ Tag Filter - รองรับการเลือกหลาย Tag (Multi-select)
        tags_param = request.query_params.get('tags') # e.g. "1,2,5"
        tag_id = request.query_params.get('tag_id')   # Single legacy

        if tags_param:
            tag_ids = [int(t) for t in tags_param.split(',') if t.isdigit()]
            if tag_ids:
                products = products.filter(tags__id__in=tag_ids).distinct()
        elif tag_id:
             products = products.filter(tags__id=tag_id).distinct()
        
        tag_slug = request.query_params.get('tag')
        if tag_slug and not tags_param and not tag_id:
            # ใช้ Slug หรือ Name ในการค้นหา
            products = products.filter(Q(tags__slug=tag_slug) | Q(tags__name__iexact=tag_slug)).distinct()

        # ✅ In Stock Filter
        in_stock = request.query_params.get('in_stock')
        if in_stock == 'true':
            products = products.filter(stock__gt=0)

        if search:
            products = products.filter(title__icontains=search)

        paginator = PageNumberPagination()
        paginator.page_size = 50
        result_page = paginator.paginate_queryset(products, request)
        
        data = []
        for p in result_page:
            try:
                thumbnail_url = ""
                if p.thumbnail:
                    thumbnail_url = p.thumbnail.url
                
                data.append({
                    "id": p.id,
                    "title": p.title,
                    "category": p.category.name if p.category else "", # ✅ Fix serialization
                    "price": p.price,
                    "stock": p.stock,
                    "rating": p.rating,
                    "thumbnail": thumbnail_url,
                })
            except Exception:
                continue
            
        return paginator.get_paginated_response(data)
            
    except Exception as e:
        err_msg = f"Cannot fetch products: {str(e)}\n{traceback.format_exc()}"
        logger.error(err_msg)
        return Response({"error": str(e)}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_all_products_admin_api(request):
    if request.user.role not in ['admin', 'super_admin', 'seller']:
        return Response({"error": "Unauthorized"}, status=403)
    
    # ✅ Fix: Show only Active products (Hide Deleted)
    # ✅ Fix: Annotate sales_count for Automation
    # ✅ Fix: Prefetch tags for Edit Tag Modal
    products = Product.objects.filter(is_active=True).annotate(
        sales_count=Sum('order_items__quantity')
    ).prefetch_related('tags').order_by('-id')
    
    data = []
    for p in products:
        data.append({
            "id": p.id,
            "title": p.title,
            "price": p.price,
            "stock": p.stock,
            "sales_count": p.sales_count or 0, # ✅ Return sales_count
            "category": p.category.name if p.category else "-",
            "tags": [{"id": t.id, "name": t.name} for t in p.tags.all()], # ✅ Return Tags
            "is_active": p.is_active,
            "thumbnail": p.thumbnail.url if p.thumbnail else "/placeholder.png"
        })
    return Response(data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def manage_user_role(request):
    # อนุญาตให้ admin จัดการได้
    if request.user.role != 'admin':
        return Response({"error": "Unauthorized"}, status=403)
    
    user_id = request.data.get('user_id')
    new_role = request.data.get('new_role') # รับค่า role ใหม่โดยตรง

    try:
        target_user = User.objects.get(id=user_id)
        
        # ป้องกันการเปลี่ยน role ตัวเอง
        if target_user.id == request.user.id:
             return Response({"error": "ไม่สามารถเปลี่ยนสถานะตัวเองได้"}, status=400)

        # ตรวจสอบค่า role ที่ส่งมา
        valid_roles = dict(User.Role.choices).keys()
        if new_role not in valid_roles:
            return Response({"error": "Invalid role"}, status=400)

        # บันทึก Role ใหม่
        target_user.role = new_role
        

        target_user.save()
        
        AdminLog.objects.create(admin=request.user, action=f"Changed role of {target_user.username} to {new_role}")
        
        return Response({"message": f"เปลี่ยนสิทธิ์เป็น {new_role} สำเร็จ"})
        
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=404)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_user_api(request, user_id):
    # อนุญาตให้ admin ลบผู้ใช้ได้
    if request.user.role != 'admin':
        return Response({"error": "Unauthorized"}, status=403)

    try:
        target_user = User.objects.get(id=user_id)
        
        # ป้องกันการลบตัวเอง
        if target_user.id == request.user.id:
             return Response({"error": "ไม่สามารถลบบัญชีตัวเองได้"}, status=400)

        # (Optional) ป้องกันการลบ Super Admin อื่น ถ้าเราไม่ใช่ Super Admin (แต่ในที่นี้ถือว่า Admin ร้านจัดการได้หมดถ้าระดับเดียวกัน)
        if target_user.role == 'admin' and request.user.role != 'admin':
             return Response({"error": "ไม่สามารถลบ Admin (Super User เดิม) ได้"}, status=403)

        username = target_user.username
        target_user.delete()
        
        AdminLog.objects.create(admin=request.user, action=f"Deleted user: {username}")
        
        return Response({"message": f"ลบผู้ใช้ {username} สำเร็จ"})
        
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=404)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def admin_update_user_api(request, user_id):
    # อนุญาตให้ seller, admin, super_admin
    if request.user.role not in ['seller', 'admin', 'super_admin']:
        return Response({"error": "Unauthorized"}, status=403)

    try:
        try:
            target_user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({"error": "User not found in DB"}, status=404)
        
        data = request.data

        # 1. ตรวจสอบ Username ซ้ำ (ถ้าเปลี่ยน)
        new_username = data.get('username')
        if new_username and new_username != target_user.username:
            if User.objects.filter(username=new_username).exists():
                return Response({"error": "Username นี้มีผู้ใช้แล้ว"}, status=400)
            target_user.username = new_username

        # 2. ตรวจสอบ Email ซ้ำ (ถ้าเปลี่ยน)
        new_email = data.get('email')
        if new_email and new_email != target_user.email:
            if User.objects.filter(email=new_email).exists():
                return Response({"error": "Email นี้มีผู้ใช้แล้ว"}, status=400)
            target_user.email = new_email

        # 3. อัปเดต Role
        new_role = data.get('role')
        if new_role:
             # ป้องกันการเปลี่ยน role ตัวเอง (ยกเว้น migrate จาก super_admin -> admin)
            if target_user.id == request.user.id and new_role != target_user.role:
                 if not (target_user.role == 'super_admin' and new_role == 'admin'):
                     return Response({"error": "ไม่สามารถเปลี่ยน Role ตัวเองได้ที่นี่"}, status=400)
            
            valid_roles = dict(User.Role.choices).keys()
            if new_role in valid_roles:
                target_user.role = new_role


        # 4. อัปเดตข้อมูลอื่นๆ
        if 'first_name' in data: target_user.first_name = data['first_name']
        if 'last_name' in data: target_user.last_name = data['last_name']
        if 'phone' in data: target_user.phone = data['phone']
        if 'address' in data: target_user.address = data['address']
        if 'is_active' in data: 
            # Prevent self-deactivation
            if target_user.id == request.user.id and str(data['is_active']).lower() == 'false':
                 return Response({"error": "ไม่สามารถปิดการใช้งานบัญชีตัวเองได้"}, status=400)
            target_user.is_active = data['is_active']
        if 'avatar' in request.FILES: target_user.image = request.FILES['avatar'] # Fixed field name to image

        target_user.save()

        AdminLog.objects.create(admin=request.user, action=f"Updated user details: {target_user.username}")
        return Response({"message": "อัปเดตข้อมูลผู้ใช้สำเร็จ"})

    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=500)

# ==========================================
# 🛒 Public API (สินค้าหน้าบ้าน)
# ==========================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reply_review_api(request, review_id):
    try:
        review = Review.objects.get(id=review_id)
        # Check permission: Admin or Seller
        if not (request.user.role in ['admin', 'super_admin', 'seller']):
             return Response({"error": "Unauthorized"}, status=403)

        review.reply_comment = request.data.get('reply_comment')
        review.reply_timestamp = timezone.now()
        review.save()
        
        # Log action
        if request.user.role != 'seller': # Optional: log for admin
             AdminLog.objects.create(admin=request.user, action=f"Replied to review {review_id}", details=f"Reply: {review.reply_comment[:50]}...")

        return Response({"message": "Reply added", "reply_comment": review.reply_comment, "reply_date": review.reply_timestamp.strftime("%d/%m/%Y %H:%M")})
    except Review.DoesNotExist:
        return Response({"error": "Review not found"}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['GET'])
@permission_classes([AllowAny])
def product_detail_api(request, product_id):
    print(f"DEBUG: product_detail_api called with id={product_id}")
    try:
        product = Product.objects.get(id=product_id) # Renamed 'p' to 'product'
        gallery = []
        try:
            gallery = [{"id": img.id, "image": img.image_url.url} for img in product.images.all()] # Used image_url field name
        except Exception as e:
            print(f"Error processing gallery images for product {product_id}: {e}")
            traceback.print_exc()
        
        reviews = []
        try:
            for r in product.reviews.all():
                reviews.append({
                    "id": r.id,
                    "user": r.user.username if r.user else "Anonymous",
                    "rating": r.rating,
                    "comment": r.comment,
                    "date": r.created_at.strftime("%d/%m/%Y") if r.created_at else "",
                    "image": r.image.url if r.image else None,  # ✅ เพิ่มรูปภาพรีวิว
                    # ✅ Return Reply Data
                    "reply_comment": r.reply_comment,
                    "reply_date": r.reply_timestamp.strftime("%d/%m/%Y %H:%M") if r.reply_timestamp else ""
                })
        except Exception as e:
            print(f"Error processing reviews for product {product_id}: {e}")
            traceback.print_exc()

        # ✅ Next/Prev IDs for navigation
        next_product = Product.objects.filter(id__gt=product.id, is_active=True).order_by('id').first()
        prev_product = Product.objects.filter(id__lt=product.id, is_active=True).order_by('-id').first()

        data = {
        "id": product.id,
        "title": product.title,
        "description": product.description,
        "price": product.price,
        "original_price": product.original_price,
        "stock": product.stock,
        "brand": product.brand,
        "category": product.category.name if product.category else "",
        "tags": [{"id": t.id, "name": t.name, "icon": t.icon, "color": t.color, "slug": t.slug} for t in product.tags.all()], # ✅ Detailed Tags for URL/Badges
        "sku": product.sku,
        "weight": product.weight,
        "dimensions": {
            "width": product.width,
            "height": product.height,
            "depth": product.depth
        },
        "thumbnail": product.thumbnail.url if product.thumbnail else "",
        "rating": product.rating,
        "images": gallery,
        "reviews": reviews, 
        "seller": {
            "id": product.admin.id,
            "username": product.admin.username,
            "shop_name": product.admin.first_name
        } if product.admin else None,
        "next_id": next_product.id if next_product else None,


        }

        # ✅ Check Flash Sale Logic (Manual Injection)
        now = timezone.now()
        active_fs = FlashSaleProduct.objects.filter(
            product=product,
            flash_sale__start_time__lte=now,
            flash_sale__end_time__gte=now,
            flash_sale__is_active=True
        ).first()

        if active_fs and active_fs.sold_count < active_fs.quantity_limit:
            data["flash_sale_info"] = {
                'id': active_fs.flash_sale.id,
                'sale_price': active_fs.sale_price,
                'end_time': active_fs.flash_sale.end_time,
                'quantity_limit': active_fs.quantity_limit,
                'sold_count': active_fs.sold_count
            }
        else:
            data["flash_sale_info"] = None

        return Response(data)
    except Product.DoesNotExist:
        # Debugging: Print all existing IDs to see what's actually in the DB
        all_ids = list(Product.objects.values_list('id', flat=True))
        print(f"DEBUG: Product {product_id} not found. Existing IDs: {all_ids}")
        return Response({"error": f"Not found. Encoded ID: {product_id}. Available: {len(all_ids)} products"}, status=404)
    except Exception as e:
        traceback.print_exc()
        return Response({"error": str(e)}, status=500)
import json
from urllib.request import urlopen

@api_view(['GET'])
@permission_classes([AllowAny])
def categories_api(request):
    # 1. หมวดหมู่มาตรฐาน (Fallback)
    default_categories = ["Furniture", "Decor", "Lighting", "Bedding", "Electronics", "Clothing"]

    # 2. ดึงจาก DummyJSON
    try:
        url = "https://dummyjson.com/products/category-list"
        with urlopen(url, timeout=5) as response:
            data = json.loads(response.read().decode())
            if isinstance(data, list):
                # Save/Update categories in DB
                for cat_name in data:
                    # Convert to Title Case for consistency
                    pretty_name = cat_name.title()
                    Category.objects.get_or_create(name=pretty_name)
                    
    except Exception as e:
        print(f"⚠️ Error fetching DummyJSON categories: {e}")
        # If fetch fails, ensure defaults exist
        for cat_name in default_categories:
            Category.objects.get_or_create(name=cat_name)

    # 3. Return only categories with active products (Hide empty categories)
    all_categories = Category.objects.filter(products__is_active=True).distinct().order_by('name').values_list('name', flat=True)
    
    return Response({"categories": list(all_categories)})


@api_view(['GET'])
@permission_classes([AllowAny])
def brands_api(request):
    """
    Get all unique brands.
    Optional query param: ?category=... to filter brands by category logic.
    """
    category = request.query_params.get('category')
    
    products = Product.objects.filter(is_active=True).exclude(brand__isnull=True).exclude(brand="")
    
    if category and category != "ทั้งหมด":
        products = products.filter(category__name=category)
        
    brands = products.values_list('brand', flat=True).distinct().order_by('brand')
    
    # Return list, sorted alphabetically usually good
    return Response({"brands": ["ทั้งหมด"] + list(brands)})

# ==========================================
# 📝 Auth & Profile
# ==========================================

@api_view(['POST'])
@permission_classes([AllowAny])
def register_api(request):
    # รับข้อมูลทั้งหมด
    data = request.data 
    
    username = data.get('username')
    password = data.get('password')
    email = data.get('email')
    phone = data.get('phone', '') 
    first_name = data.get('first_name', '') # รับชื่อจริง
    last_name = data.get('last_name', '')   # รับนามสกุล
    
    if User.objects.filter(username=username).exists():
        return Response({"error": "ชื่อผู้ใช้นี้ถูกใช้งานแล้ว (Username already exists)"}, status=400)

    # ✅ เพิ่มการตรวจสอบ Email ซ้ำ
    if User.objects.filter(email=email).exists():
        return Response({"error": "อีเมลนี้ถูกใช้งานแล้ว (Email already exists)"}, status=400)

    try:
        # สร้าง User หลัก
        user = User.objects.create_user(username=username, password=password, email=email)
        user.role = 'new_user' 
        user.phone = phone
        user.first_name = first_name # บันทึกชื่อจริง
        user.last_name = last_name   # บันทึกนามสกุล
        
        # ถ้ารูปถูกส่งมาด้วย ให้บันทึกลง profile -> user model
        if 'avatar' in request.FILES:
            user.image = request.FILES['avatar']
        
        user.save()
            
        return Response({"message": "Registered successfully"})
    except Exception as e:
        import traceback
        traceback.print_exc() # 🖨️ Print Error to Terminal for debugging
        return Response({"error": f"เกิดข้อผิดพลาด: {str(e)}"}, status=400)

@api_view(['GET'])
@permission_classes([AllowAny])
def check_username_api(request):
    username = request.query_params.get('username', None)
    if not username:
        return Response({"error": "No username provided"}, status=400)
    
    is_taken = User.objects.filter(username=username).exists()
    return Response({"available": not is_taken})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_system_user(request):
    """
    API สำหรับ Admin/SuperUser สร้าง System User ใหม่
    """
    # ตรวจสอบสิทธิ์: ต้องเป็น Admin เท่านั้น
    if request.user.role != 'admin':
        return Response({"error": "Unauthorized: คุณไม่มีสิทธิ์สร้างผู้ใช้งาน"}, status=403)
    
    data = request.data
    username = data.get('username')
    password = data.get('password')
    email = data.get('email')
    role = data.get('role')
    
    if not all([username, password, email, role]):
        return Response({"error": "กรุณากรอกข้อมูลให้ครบถ้วน (Username, Password, Email, Role)"}, status=400)
        
    if User.objects.filter(username=username).exists():
        return Response({"error": "ชื่อผู้ใช้นี้ถูกใช้งานแล้ว"}, status=400)
        
    if User.objects.filter(email=email).exists():
        return Response({"error": "อีเมลนี้ถูกใช้งานแล้ว"}, status=400)
        
    try:
        user = User.objects.create_user(username=username, password=password, email=email)
        user.role = role
        user.first_name = data.get('first_name', '')
        user.last_name = data.get('last_name', '')
        user.phone = data.get('phone', '')
        user.address = data.get('address', '')
        user.is_active = True
        
        if 'avatar' in request.FILES:
            user.image = request.FILES['avatar']
            
        user.save()
        
        AdminLog.objects.create(admin=request.user, action=f"Created new user: {username} (Role: {role})")
        
        return Response({"message": f"สร้างผู้ใช้งาน {username} สำเร็จ", "id": user.id}, status=201)
        
    except Exception as e:
        return Response({"error": f"เกิดข้อผิดพลาด: {str(e)}"}, status=500)


@api_view(['POST'])
@permission_classes([AllowAny])
def login_api(request):
    username = request.data.get('username')
    password = request.data.get('password')
    
    if not username or not password:
         return Response({"error": "Please provide both username and password"}, status=400)

    # ✅ Support Email Login
    if '@' in username:
        try:
            user_obj = User.objects.get(email=username)
            username = user_obj.username # Normalize to username for authenticate
        except User.DoesNotExist:
            pass # Let authenticate fail naturally

    user = authenticate(username=username, password=password)
    
    if user:
        if not user.is_active:
             return Response({"error": "Account disabled"}, status=403)
             
        token, _ = Token.objects.get_or_create(user=user)
        print(f"DEBUG: Login Success for {user.username}. Token: {token.key}")
        
        # Determine Role logic (Helper for frontend)
        role = user.role
        if user.is_superuser: role = 'admin'

        return Response({
            "token": token.key,
            "id": user.id,
            "username": user.username,
            "role": role, 
            "role_code": user.role,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "email": user.email,
            "phone": user.phone,
            "address": user.address,
            "avatar": user.image.url if user.image else ""
        })
    
    return Response({"error": "ชื่อผู้ใช้งานหรือรหัสผ่าน ไม่ถูกต้อง"}, status=401)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_api(request):
    # Token deletion not standard in session auth but if using token auth:
    if hasattr(request.user, 'auth_token'):
        request.user.auth_token.delete()
    return Response({"message": "Logged out"})

@api_view(['GET', 'PUT']) 
@permission_classes([IsAuthenticated])
def user_profile_api(request):
    user = request.user

    # 🟢 กรณีดึงข้อมูล (GET)
    if request.method == 'GET':
        return Response({
            "id": user.id,
            "username": user.username,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "email": user.email,
            "phone": user.phone,
            "address": user.address, 
            "role": user.get_role_display(),
            "role_code": user.role,
            "avatar": user.image.url if user.image else "" # Use image field
        })
    
    # 🟠 กรณีบันทึกแก้ไข (PUT)
    elif request.method == 'PUT':
        data = request.data
        
        try:
            # 1. อัปเดตข้อมูล User หลัก (username, email)
            # ✅ Validations for Duplicates
            if 'username' in data and data['username'] != user.username:
                if User.objects.filter(username=data['username']).exists():
                    return Response({"error": "ชื่อผู้ใช้นี้มีผู้ใช้งานแล้ว (Username unavailable)"}, status=400)
                user.username = data['username']

            if 'email' in data and data['email'] != user.email:
                if User.objects.filter(email=data['email']).exists():
                    return Response({"error": "อีเมลนี้มีผู้ใช้งานแล้ว (Email unavailable)"}, status=400)
                user.email = data['email']

            if 'first_name' in data: user.first_name = data['first_name']
            if 'last_name' in data: user.last_name = data['last_name']

            # 2. อัปเดตข้อมูล Profile fields (phone, address, image)
            if 'phone' in data: user.phone = data['phone']
            if 'address' in data: user.address = data['address']
            if 'avatar' in request.FILES: user.image = request.FILES['avatar']
            
            user.save()

            # Return full user object to update frontend immediately
            return Response({
                "message": "บันทึกข้อมูลสำเร็จ",
                "id": user.id,
                "username": user.username,
                "role": user.get_role_display(),
                "role_code": user.role,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "email": user.email,
                "phone": user.phone,
                "address": user.address,
                "avatar": user.image.url if user.image else ""
            })
        except Exception as e:
            return Response({"error": f"เกิดข้อผิดพลาด: {str(e)}"}, status=400)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_password_api(request):
    try:
        user = request.user
        old_password = request.data.get('old_password')
        new_password = request.data.get('new_password')
        
        if not old_password or not new_password:
            return Response({"error": "กรุณากรอกรหัสผ่านเดิมและรหัสผ่านใหม่"}, status=400)
        
        if not user.check_password(old_password):
            return Response({"error": "รหัสผ่านเดิมไม่ถูกต้อง"}, status=400)
        
        if len(new_password) < 6:
            return Response({"error": "รหัสผ่านใหม่ต้องมีความยาวอย่างน้อย 6 ตัวอักษร"}, status=400)

        user.set_password(new_password)
        user.save()
        
        # Keep user logged in
        # update_session_auth_hash(request, user)  
        
        return Response({"message": "เปลี่ยนรหัสผ่านสำเร็จ"})
    except Exception as e:
        return Response({"error": f"เกิดข้อผิดพลาด: {str(e)}"}, status=400)

# ==========================================
# 📦 Order & Stats
# ==========================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_order(request):
    # ✅ 1. รับข้อมูล
    cart_items = request.data.get('items') or request.data.get('cart_items', [])
    customer_data = request.data.get('customer', {}) 
    coupon_code = request.data.get('couponCode') # ✅ Coupon Code
    
    # 🛡️ Inline Validation (Fail Fast)
    # Combine data for validation
    validation_payload = {
        "items": cart_items,
        "name": customer_data.get('name'),
        "tel": customer_data.get('phone') or customer_data.get('tel'),
        "address": customer_data.get('address'),
        "province": customer_data.get('province')
    }
    
    try:
        validate_order_data(validation_payload)
    except InlineValidationError as e:
        # 🚫 Return 422 directly without logging (as designed)
        return Response(e.detail, status=422)

    if not cart_items: 
        return Response({"error": "ไม่พบรายการสินค้า (Empty cart)"}, status=400)

    try:
        with transaction.atomic():
            total_price = 0
            order_items_to_create = []
            has_flash_sale_item = False
            now = timezone.now()
            
            # ✅ 2. Loop Process Items
            for item in cart_items:
                try:
                    p = Product.objects.select_for_update().get(id=item['id'])
                except Product.DoesNotExist:
                    raise ValueError(f"สินค้า ID {item['id']} ไม่พบในระบบ")

                qty = int(item['quantity'])

                # ⚡ Check Active Flash Sale
                active_fs = FlashSaleProduct.objects.filter(
                    product=p,
                    flash_sale__start_time__lte=now,
                    flash_sale__end_time__gte=now,
                    flash_sale__is_active=True
                ).select_for_update().first()
                
                # Default Price
                final_price = p.price
                
                # Logic: If Flash Sale available AND stock available logic
                if active_fs:
                    # Check flash sale limits
                    if active_fs.sold_count < active_fs.quantity_limit:
                         # Use Flash Sale Price
                         final_price = active_fs.sale_price
                         has_flash_sale_item = True
                         
                         # Check if enough flash sale quota for this order?
                         # For simplicity: If qty > remaining flash sale stock, we might split or reject.
                         # User requirement: "Prevent oversell".
                         # Let's reject if FS stock not enough for full quantity or mixed.
                         fs_remaining = active_fs.quantity_limit - active_fs.sold_count
                         if qty > fs_remaining:
                             raise ValueError(f"สินค้า '{p.title}' เหลือสิทธิ์ Flash Sale เพียง {fs_remaining} ชิ้น")
                             
                         active_fs.sold_count += qty
                         active_fs.save()
                    
                # Check Main Product Stock
                if p.stock < qty: 
                    raise ValueError(f"สินค้า '{p.title}' สินค้าหมดหรือมีไม่พอ (เหลือ {p.stock})")
                
                # Deduct Main Stock
                p.stock -= qty
                p.save()
                
                item_total = final_price * qty
                total_price += item_total
                
                order_items_to_create.append({
                    "product": p,
                    "quantity": qty,
                    "price": final_price
                })

            # ✅ 3. Apply Coupon
            item_subtotal_val = total_price # Rename for clarity
            shipping_cost = 50 # Default Shipping Cost (Flat Rate)
            
            # Recalculate Total (Subtotal + Shipping)
            grand_total = item_subtotal_val + shipping_cost
            
            discount_amount = 0
            coupon = None
            
            if coupon_code:
                # 🔒 Lock Coupon Row to prevent Race Condition
                try:
                     coupon_locked = Coupon.objects.select_for_update().get(code=coupon_code)
                except Coupon.DoesNotExist:
                     raise ValueError("รหัสคูปองไม่ถูกต้อง")

                # ✅ Refactored Logic using CouponService
                # Pass shipping_cost to validate if needed (future)
                is_valid, msg, coupon = CouponService.validate_coupon(
                    user=request.user, 
                    coupon_code=coupon_code, 
                    cart_total=item_subtotal_val, # Validate against Item Subtotal
                    cart_items=cart_items
                )
                
                if not is_valid:
                    raise ValueError(msg)

                # ⚡ Flash Sale Stackability Check
                if has_flash_sale_item and not coupon.is_stackable_with_flash_sale:
                     raise ValueError("คูปองนี้ไม่สามารถใช้ร่วมกับสินค้า Flash Sale ได้")

                # ✅ Calculate Discount (Pass Shipping Cost)
                discount_amount = CouponService.calculate_discount(coupon, item_subtotal_val, shipping_cost)
                
                # Prevent negative total
                # Max discount cannot exceed (Subtotal + Shipping) effectively
                # But logic says: Free Shipping = 50. 
                # Percent/Fixed = based on Subtotal usually.
                # Let's ensure discount doesn't exceed Grand Total
                discount_amount = min(discount_amount, grand_total)
                
                # Update Usage
                coupon.used_count = F('used_count') + 1
                coupon.save()
                
                grand_total -= discount_amount

            # ✅ 4. Create Order
            order = Order.objects.create(
                user=request.user,
                customer_name=customer_data.get('name', request.user.first_name or request.user.username),

                customer_tel=customer_data.get('phone', customer_data.get('tel', request.user.phone)), 
                customer_email=customer_data.get('email', request.user.email),
                shipping_address=customer_data.get('address', request.user.address), 
                shipping_province=customer_data.get('province'), 
                payment_method=request.data.get('paymentMethod', 'Transfer'),
                
                item_subtotal=item_subtotal_val,
                shipping_cost=shipping_cost,
                total_price=grand_total,
                
                discount_amount=discount_amount, # ✅ Save Discount
                coupon=coupon, # ✅ Save Coupon
                status='Pending'
            )
            
            # ✅ 5. Create Order Items
            for item_data in order_items_to_create:
                OrderItem.objects.create(
                    order=order, 
                    product=item_data['product'], 
                    quantity=item_data['quantity'], 
                    price_at_purchase=item_data['price'] 
                )

            # ✅ 6. Update User Role (New User -> Customer)
            if request.user.role == 'new_user':
                request.user.role = 'customer'
                request.user.save()

        return Response({"message": "สั่งซื้อสำเร็จ", "order_id": order.id, "total": total_price}, status=201)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({"error": str(e)}, status=400)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_orders_api(request):
    # ✅ Optimization: Prefetch items and products to prevent N+1 queries
    orders = Order.objects.filter(user=request.user).prefetch_related('items__product').order_by('-created_at')
    data = []
    for o in orders:
        try:
            items = []
            for i in o.items.all(): # related_name='items' in OrderItem
                if not i.product: continue
                thumb = i.product.thumbnail.url if i.product.thumbnail else ""
                items.append({"title": i.product.title, "quantity": i.quantity, "price": i.price_at_purchase, "thumbnail": thumb})
            
            date_str = o.created_at.strftime("%d/%m/%Y") if o.created_at else "-"
            
            # ✅ Generate QR Payload for Pending orders
            qr_payload = ""
            if o.status == 'Pending' and not o.slip_image:
                 qr_payload = generate_promptpay_payload(o.total_price)
                 
            data.append({
                "id": o.id, "date": date_str, 
                "total_price": o.total_price, "status": o.status, "items": items,
                "has_slip": bool(o.slip_image),
                "promptpay_payload": qr_payload, # ✅ For Frontend QR
                "province": o.shipping_province # ✅ Return Province
            })
        except Exception as e:
            print(f"Error processing order {o.id}: {e}")
            continue
    return Response(data)

# ✅ Alias for Checkout API
checkout_api = create_order

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_slip_api(request, order_id):
    try:
        order = Order.objects.get(id=order_id, user=request.user)
        
        file = request.FILES.get('slip_image')
        if not file:
            return Response({"error": "ไม่พบไฟล์สลิป"}, status=400)
            
        # Update Slip Data
        order.slip_image = file
        order.transfer_date = request.data.get('transfer_date')
        order.transfer_amount = request.data.get('transfer_amount', 0)
        order.bank_name = request.data.get('bank_name', '')
        order.transfer_account_number = request.data.get('transfer_account_number', '')
        order.status = 'Pending' # Confirm pending status
        order.save()
        
        return Response({"message": "อัพโหลดสลิปเรียบร้อยแล้ว"})
    except Order.DoesNotExist:
        return Response({"error": "ไม่พบคำสั่งซื้อหรือคุณไม่มีสิทธิ์"}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=400)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def confirm_received_api(request, order_id):
    try:
        # ✅ Allowed for Order Owner
        order = Order.objects.get(id=order_id, user=request.user)
        if order.status == 'Shipped':
            order.status = 'Completed'
            order.save()
            return Response({"message": "Confirmed received"})
        return Response({"error": "Invalid status - Must be Shipped"}, status=400)
    except Order.DoesNotExist:
        return Response({"error": "Order not found or not authorized"}, status=404)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_admin_stats(request):
    if request.user.role not in ['seller', 'admin', 'super_admin']: return Response(status=403)
    
    # 1. Basic Stats (Global)
    # Include all confirmed sales/money-in statuses
    VALID_SALES_STATUSES = ['Paid', 'Processing', 'Shipped', 'Completed']
    total_sales = Order.objects.filter(status__in=VALID_SALES_STATUSES).aggregate(Sum('total_price'))['total_price__sum'] or 0
    total_orders = Order.objects.count()
    total_products = Product.objects.count()
    total_users = User.objects.count()

    # 2. Sales Chart Data (Dynamic Period)
    period = request.GET.get('period', 'daily')
    date_param = request.GET.get('date')
    
    if date_param:
        try:
             # Handle Javascript ISO format
            if 'T' in date_param:
                target_date = datetime.fromisoformat(date_param.replace("Z", "+00:00")).date()
            else:
                target_date = datetime.strptime(date_param, "%Y-%m-%d").date()
        except:
            target_date = timezone.now().date()
    else:
        target_date = timezone.now().date()

    # Define Filter based on Period
    # sales_filter = {'status': 'Paid'} -> CHANGED
    sales_filter = {'status__in': VALID_SALES_STATUSES}
    orders_filter = {}

    if period == 'daily':
        sales_filter['created_at__date'] = target_date
        orders_filter['created_at__date'] = target_date
    elif period == 'monthly':
        sales_filter['created_at__year'] = target_date.year
        sales_filter['created_at__month'] = target_date.month
        orders_filter['created_at__year'] = target_date.year
        orders_filter['created_at__month'] = target_date.month
    elif period == 'yearly':
        sales_filter['created_at__year'] = target_date.year
        orders_filter['created_at__year'] = target_date.year

    # 1. Basic Stats (Filtered)
    # ...
    
    # 1. Basic Stats (Filtered by Date/Period)
    # Note: If category_id is needed globally, it should be applied here too.
    cat_id = request.GET.get('category_id')
    
    # Apply Category Filter if present
    if cat_id and cat_id != 'all':
        sales_filter['items__product__category_id'] = cat_id
        orders_filter['items__product__category_id'] = cat_id
        
        # When filtering by category, total_sales should be sum of ITEM sales, not Order total
        # But for simplicity/performance in this view, we might filter Orders that contain the category
        # For accurate revenue:
        total_sales = OrderItem.objects.filter(
            order__status__in=VALID_SALES_STATUSES,
            product__category_id=cat_id,
            **{f"order__{k}": v for k, v in sales_filter.items() if k != 'items__product__category_id' and k != 'status__in'}
        ).aggregate(
            total=Sum(F('price_at_purchase') * F('quantity'))
        )['total'] or 0
        
        total_orders = Order.objects.filter(**orders_filter).distinct().count()
    else:
        total_sales = Order.objects.filter(**sales_filter).aggregate(Sum('total_price'))['total_price__sum'] or 0
        total_orders = Order.objects.filter(**orders_filter).count()

    # ...

    # --- Safely Calculate Sales Data ---
    sales_data = [] # Reset/Init
    
    try:
        if period == 'daily':
            # Last 7 Days from target_date
            for i in range(6, -1, -1):
                 d = target_date - timedelta(days=i)
                 # Use status__in logic
                 sales = Order.objects.filter(status__in=VALID_SALES_STATUSES, created_at__date=d).aggregate(Sum('total_price'))['total_price__sum'] or 0
                 sales_data.append({"name": d.strftime("%d/%m"), "sales": sales})

        elif period == 'monthly':
            # Show all days in target_date's month
            _, num_days = calendar.monthrange(target_date.year, target_date.month)
            print(f"DEBUG: Monthly {target_date.month}/{target_date.year} has {num_days} days")
            for day in range(1, num_days + 1):
                 d = target_date.replace(day=day)
                 sales = Order.objects.filter(status__in=VALID_SALES_STATUSES, created_at__date=d).aggregate(Sum('total_price'))['total_price__sum'] or 0
                 sales_data.append({"name": str(day), "sales": sales})

        elif period == 'yearly':
            # Show all months in target_date's year
            print(f"DEBUG: Yearly {target_date.year}")
            months_th = ["ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.", "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]
            for month in range(1, 13):
                 sales = Order.objects.filter(status__in=VALID_SALES_STATUSES, created_at__year=target_date.year, created_at__month=month).aggregate(Sum('total_price'))['total_price__sum'] or 0
                 sales_data.append({"name": months_th[month-1], "sales": sales})
    except Exception as e:
        print(f"DEBUG: Sales Data Calculation Error: {e}")
        sales_data = []

    # --- Safely Calculate Extra Stats ---
    best_sellers_data = []
    pie_data = []
    low_stock_data = []
    logs_data = []
    pending_orders = Order.objects.filter(status='Pending').count()

    try:
        # 2. Best Sellers (Top 5 Products) - Filtered
        # Group by Product ID first to match specific products
        best_sellers_query = OrderItem.objects.filter(order__in=Order.objects.filter(**orders_filter))\
            .values('product__id')\
            .annotate(total_qty=Sum('quantity'))\
            .order_by('-total_qty')[:5]
        
        # Fetch actual Product objects for details
        product_ids = [item['product__id'] for item in best_sellers_query]
        products = Product.objects.filter(id__in=product_ids).in_bulk()

        for item in best_sellers_query:
            product = products.get(item['product__id'])
            if product:
                best_sellers_data.append({
                    "name": product.title,
                    "sales": item['total_qty'],
                    "price": float(product.price), # Ensure float
                    "thumbnail": product.thumbnail.url if product.thumbnail else None
                })

    except Exception as e:
        print(f"DEBUG: Best Sellers Error: {e}")

    try:
        # 3. Low Stock
        low_stock_products = Product.objects.filter(stock__lte=10, is_active=True) # Fetch objects
        low_stock_data = [{
            "id": p.id,
            "title": p.title,
            "stock": p.stock,
            "thumbnail": p.thumbnail.url if p.thumbnail else None
        } for p in low_stock_products]
    except Exception as e:
        print(f"DEBUG: Low Stock Error: {e}")

    try:
        # 4. Pie Chart
        # Use VALID_SALES_STATUSES
        category_stats = Order.objects.filter(**orders_filter, status__in=VALID_SALES_STATUSES)\
            .values('items__product__category')\
            .annotate(value=Sum('items__price_at_purchase'))\
            .order_by('-value')
        
        for c in category_stats:
            cat_name = c['items__product__category'] or "Uncategorized"
            if c['value']:
                pie_data.append({"name": cat_name, "value": c['value']})
    except Exception as e:
        print(f"DEBUG: Pie Chart Error: {e}")

    try:
        # 5. Logs
        recent_logs = AdminLog.objects.all().order_by('-timestamp')[:5]
        # Check if admin exists
        logs_data = []
        for l in recent_logs:
            user_str = "System"
            if l.admin:
                user_str = l.admin.username
            
            logs_data.append({
                "id": l.id,
                "action": l.action,
                "date": l.timestamp.strftime("%d/%m %H:%M"),
                "user": user_str,
                "target": "", 
                "time": l.timestamp.strftime("%H:%M") 
            })
    except Exception as e:
        print(f"DEBUG: Logs Error: {e}")

    try:
        # 6. Province Sales Data (Map Analytics)
        # 6. Province Sales Data (Map Analytics)
        # 6. Province Sales Data (Map Analytics)
        province_qs = Order.objects.filter(**orders_filter, status__in=VALID_SALES_STATUSES)\
            .exclude(Q(shipping_province__isnull=True) | Q(shipping_province__exact=''))

        cat_id = request.query_params.get('category_id') # Fetch here inline
        
        if cat_id and cat_id != 'all':
             province_stats = province_qs.filter(items__product__category_id=cat_id)\
                .values('shipping_province')\
                .annotate(
                    total_sales=Sum(F('items__price_at_purchase') * F('items__quantity')), 
                    total_orders=Count('id', distinct=True)
                )\
                .order_by('-total_sales')
        else:
            province_stats = province_qs\
                .values('shipping_province')\
                .annotate(total_sales=Sum('total_price'), total_orders=Count('id'))\
                .order_by('-total_sales')

        province_data = []
        for p in province_stats:
            prov_name = p['shipping_province'] or "Unknown"
            
        province_data = []
        for p in province_stats:
            prov_name = p['shipping_province'] or "Unknown"

            # Find Top 5 Products for this province
            top_products_qs = OrderItem.objects.filter(
                order__shipping_province=prov_name, 
                order__status__in=VALID_SALES_STATUSES
            ).values('product__title').annotate(
                qty=Sum('quantity'),
                sales=Sum(F('price_at_purchase') * F('quantity'))
            ).order_by('-qty')[:5]

            # Convert to list
            top_products_list = []
            for tp in top_products_qs:
                top_products_list.append({
                    "name": tp['product__title'],
                    "qty": tp['qty'], 
                    "sales": tp['sales'] or 0
                })
            
            # Use top 1 for legacy support in tooltip
            top_prod_name = top_products_list[0]['name'] if top_products_list else "-"

            province_data.append({
                "name": prov_name,
                "value": p['total_sales'],
                "order_count": p['total_orders'],
                "top_product": top_prod_name,
                "top_products_list": top_products_list
            }) 

    except Exception as e:
        print(f"DEBUG: Province Data Error: {e}")
        province_data = []
    
    # 3. Category Stats (Sales by Category)
    category_stats = []
    try:
         cat_qs = OrderItem.objects.filter(
            order__status__in=VALID_SALES_STATUSES
         )
         
         if period == 'daily':
            cat_qs = cat_qs.filter(order__created_at__date=target_date)
         elif period == 'monthly':
            cat_qs = cat_qs.filter(order__created_at__year=target_date.year, order__created_at__month=target_date.month)
         elif period == 'yearly':
            cat_qs = cat_qs.filter(order__created_at__year=target_date.year)
            
         category_stats = cat_qs.values('product__category__name').annotate(
            name=F('product__category__name'),
            value=Sum(F('price_at_purchase') * F('quantity'))
         ).order_by('-value')
    except Exception as e:
         print(f"Category Stats Error: {e}")

    return Response({
        "total_sales": total_sales,
        "total_orders": total_orders,
        "total_products": total_products,
        "total_users": total_users,
        "sales_data": sales_data,
        "best_sellers": best_sellers_data,
        "low_stock": low_stock_data,
        "pie_data": pie_data, 
        "province_data": province_data, 
        "logs": logs_data,
        "pending_orders": pending_orders,
        "category_stats": category_stats # ✅ New Data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_orders_csv(request):
    if request.user.role not in ['seller', 'admin', 'super_admin']: return Response(status=403)

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        # Create Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders Check"

        # Headers
        headers = ['รหัสสั่งซื้อ', 'วันที่', 'ลูกค้า', 'เบอร์โทร', 'รายการสินค้า', 'ยอดรวม', 'สถานะ', 'วิธีชำระ', 'สลิป']
        ws.append(headers)

        # Style Headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A4D2E", end_color="1A4D2E", fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # --- Date Filtering Logic ---
        period = request.GET.get('period', 'all')
        date_param = request.GET.get('date')
        
        orders = Order.objects.all().order_by('-created_at')

        if date_param:
            try:
                if 'T' in date_param:
                    target_date = datetime.fromisoformat(date_param.replace("Z", "+00:00")).date()
                else:
                    target_date = datetime.strptime(date_param, "%Y-%m-%d").date()
                
                if period == 'daily':
                    orders = orders.filter(created_at__date=target_date)
                elif period == 'monthly':
                    orders = orders.filter(created_at__year=target_date.year, created_at__month=target_date.month)
                elif period == 'yearly':
                    orders = orders.filter(created_at__year=target_date.year)
                    
            except Exception as e:
                print(f"Export Date Parse Error: {e}")
        
        for order in orders:
            items_str = ", ".join([f"{i.product.title} (x{i.quantity})" for i in order.items.all() if i.product])
            
            row = [
                order.id,
                order.created_at.strftime("%d/%m/%Y %H:%M"),
                order.customer_name,
                order.customer_tel,
                items_str,
                float(order.total_price), # Ensure number
                order.status,
                order.payment_method,
                "แนบแล้ว" if order.slip_image else "ยังไม่แนบ"
            ]
            ws.append(row)

        # ✅ Auto-adjust Column Width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Response
        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="orders_export.xlsx"'
        
        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({"error": str(e)}, status=500)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_admin_orders(request):
    if request.user.role not in ['seller', 'admin', 'super_admin']: return Response(status=403)
    
    orders = Order.objects.all().order_by('-created_at')

    # ✅ 1. Search (ID, Name, Tel)
    search = request.GET.get('search')
    if search:
        orders = orders.filter(Q(id__icontains=search) | Q(customer_name__icontains=search) | Q(customer_tel__icontains=search))
    
    # ✅ 2. Status Filter
    status = request.GET.get('status')
    if status and status != 'ทั้งหมด':
        orders = orders.filter(status=status)
        
    # ✅ 3. Date Filter
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        orders = orders.filter(created_at__date__gte=start_date)
    if end_date:
        orders = orders.filter(created_at__date__lte=end_date)

    # ✅ 4. Province Filter
    province = request.GET.get('province')
    if province and province != 'ทั้งหมด':
        orders = orders.filter(shipping_province=province)

    data = [{
        "id": o.id, "customer": o.customer_name, "total_price": o.total_price,
        "status": o.status, "date": o.created_at.strftime("%d/%m/%Y %H:%M"),
        "payment_method": o.payment_method,
        "slip_image": o.slip_image.url if o.slip_image else "",
        "payment_date": o.payment_date.strftime("%d/%m/%Y %H:%M") if o.payment_date else "-",
        "transfer_amount": o.transfer_amount,
        "transfer_date": o.transfer_date.strftime("%d/%m/%Y %H:%M") if o.transfer_date else "-",
        "bank_name": o.bank_name,
        "transfer_account_number": o.transfer_account_number,
        "tel": o.customer_tel,
        "province": o.shipping_province, # ✅ Add Province
        "items": [{
            "product": i.product.title if i.product else "Deleted Product",
            "quantity": i.quantity,
            "price": i.price_at_purchase
        } for i in o.items.all()]
    } for o in orders]
    
    return Response(data)


@api_view(['PUT', 'POST'])
@permission_classes([IsAuthenticated])
def update_order_status_api(request, order_id):
    # Check permissions
    if request.user.role not in ['admin', 'super_admin', 'seller']:
        return Response(status=403)
    
    try:
        order = Order.objects.get(pk=order_id)
    except Order.DoesNotExist:
        return Response({"error": "Order not found"}, status=404)
        
    new_status = request.data.get('status')
    if new_status:
        order.status = new_status
        order.save()
        
        # Log Activity if needed (optional, skipping to keep it simple and safe)
        
        return Response({"message": "Status updated", "status": new_status})
    return Response({"error": "Status required"}, status=400)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_slip(request, order_id):
    try:
        # Relaxed user check for admin/debugging or assuming user context correct
        order = Order.objects.get(pk=order_id) 
    except Order.DoesNotExist:
        return Response({"error": "Order not found"}, status=404)

    slip_image = request.FILES.get('slip_image')
    if slip_image:
        order.slip_image = slip_image
        order.transfer_amount = request.data.get('transfer_amount', 0)
        order.transfer_date = request.data.get('transfer_date')
        order.bank_name = request.data.get('bank_name', '')
        order.transfer_account_number = request.data.get('transfer_account_number', '')
        if order.status == 'pending':
            order.status = 'paid' 
        order.save()
        return Response({"message": "Slip uploaded successfully"})
    return Response({"error": "No slip image provided"}, status=400)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_update_orders_api(request):
    if request.user.role not in ['admin', 'super_admin', 'seller']:
        return Response(status=403)
    
    order_ids = request.data.get('order_ids', [])
    status = request.data.get('status')
    
    if order_ids and status:
        Order.objects.filter(id__in=order_ids).update(status=status)
        return Response({"message": "Updated"})
    return Response(status=400)


# ==========================================
# 🏷️ Tag System APIs
# ==========================================

@api_view(['GET', 'POST', 'PUT', 'DELETE', 'PATCH'])
@permission_classes([AllowAny])  # GET สามารถเข้าถึงได้ทุกคน, POST/PUT/DELETE ต้องเป็น Admin
def tag_api(request, tag_id=None):
    """
    🏷️ API สำหรับจัดการ Tags
    
    Methods:
    --------
    GET /api/tags/           # ดึง Tags ทั้งหมด
    POST /api/tags/          # สร้าง Tag ใหม่ (Admin)
    PUT /api/tags/:id/       # แก้ไข Tag (Admin)
    DELETE /api/tags/:id/    # ลบ Tag (Admin)
    """
    
    # ==========================================
    # 📖 GET - ดึงรายการ Tags ทั้งหมด
    # ==========================================
    if request.method == 'GET':
        try:
            from .models import Tag
            from .serializers import TagSerializer
            
            # ดึง Tags ทั้งหมด เรียงตามกลุ่ม และชื่อ
            tags = Tag.objects.all().order_by('group_name', 'name')
            serializer = TagSerializer(tags, many=True)
            
            return Response(serializer.data)
        except Exception as e:
            return Response({"error": str(e)}, status=500)
    
    # ==========================================
    # ➕ POST - สร้าง Tag ใหม่
    # ==========================================
    elif request.method == 'POST':
        # ✅ ตรวจสอบสิทธิ์: เฉพาะ Admin เท่านั้น
        if not request.user.is_authenticated or request.user.role not in ['admin', 'super_admin', 'seller']:
            return Response({"error": "Unauthorized"}, status=403)
        
        try:
            from .models import Tag
            from .serializers import TagSerializer
            
            # ใช้ Serializer ในการสร้างเพื่อความง่าย
            serializer = TagSerializer(data=request.data)
            if serializer.is_valid():
                # ตรวจสอบชื่อซ้ำ (Case Insensitive)
                name = serializer.validated_data.get('name')
                if Tag.objects.filter(name__iexact=name).exists():
                    return Response({"error": f"Tag '{name}' มีอยู่แล้ว"}, status=400)
                
                serializer.save()
                return Response(serializer.data, status=201)
            return Response(serializer.errors, status=400)
        except Exception as e:
            return Response({"error": str(e)}, status=500)

    # ==========================================
    # 📝 PUT/PATCH - แก้ไข Tag
    # ==========================================
    elif request.method in ['PUT', 'PATCH']:
        if not request.user.is_authenticated or request.user.role not in ['admin', 'super_admin', 'seller']:
            return Response({"error": "Unauthorized"}, status=403)
            
        if not tag_id:
            return Response({"error": "ต้องระบุ Tag ID"}, status=400)
            
        try:
            from .models import Tag
            from .serializers import TagSerializer
            
            tag = Tag.objects.get(id=tag_id)
            serializer = TagSerializer(tag, data=request.data, partial=True)
            
            if serializer.is_valid():
                # ตรวจสอบชื่อซ้ำถ้ามีการเปลี่ยนชื่อ
                new_name = serializer.validated_data.get('name')
                if new_name and new_name.lower() != tag.name.lower():
                    if Tag.objects.filter(name__iexact=new_name).exists():
                        return Response({"error": f"ชื่อ Tag '{new_name}' ถูกใช้ไปแล้ว"}, status=400)
                
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=400)
        except Tag.DoesNotExist:
            return Response({"error": "ไม่พบ Tag นี้"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)
    
    # ==========================================
    # 🗑️ DELETE - ลบ Tag
    # ==========================================
    elif request.method == 'DELETE':
        # ✅ ตรวจสอบสิทธิ์: เฉพาะ Admin เท่านั้น
        if not request.user.is_authenticated or request.user.role not in ['admin', 'super_admin', 'seller']:
            return Response({"error": "Unauthorized"}, status=403)
        
        if not tag_id:
            return Response({"error": "ต้องระบุ Tag ID"}, status=400)
        
        try:
            from .models import Tag
            
            # ค้นหา Tag
            tag = Tag.objects.get(id=tag_id)
            tag_name = tag.name
            
            # ลบ Tag (ความสัมพันธ์กับสินค้าจะถูกลบอัตโนมัติ)
            tag.delete()
            
            return Response({"message": f"ลบ Tag '{tag_name}' สำเร็จ"})
        except Tag.DoesNotExist:
            return Response({"error": "ไม่พบ Tag นี้"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def product_tags_api(request, product_id):
    """
    🔗 API สำหรับกำหนด Tags ให้กับสินค้า
    
    ใช้สำหรับ:
    - บันทึก/อัปเดต Tags ของสินค้าจาก Admin Panel
    
    Method:
    -------
    POST /api/products/:id/tags/
    
    Payload:
    --------
    {
        "tag_ids": [1, 3, 5]  # รายการ Tag IDs ที่ต้องการกำหนด
    }
    """
    
    # ✅ ตรวจสอบสิทธิ์: เฉพาะ Admin เท่านั้น
    if request.user.role not in ['admin', 'super_admin', 'seller']:
        return Response({"error": "Unauthorized"}, status=403)
    
    try:
        from .models import Product, Tag
        
        # ค้นหาสินค้า
        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return Response({"error": "ไม่พบสินค้านี้"}, status=404)
        
        # รับ Tag IDs จาก request
        tag_ids = request.data.get('tag_ids', [])
        
        # ตรวจสอบว่าเป็น list
        if not isinstance(tag_ids, list):
            return Response({"error": "tag_ids ต้องเป็น array"}, status=400)
        
        # ลบ Tags เก่าทั้งหมด
        product.tags.clear()
        
        # เพิ่ม Tags ใหม่
        for tag_id in tag_ids:
            try:
                tag = Tag.objects.get(id=tag_id)
                product.tags.add(tag)
            except Tag.DoesNotExist:
                # ถ้า Tag ไม่เจอ ข้ามไป (ไม่ error)
                continue
        
        # ส่งผลลัพธ์กลับ
        from .serializers import TagSerializer
        updated_tags = TagSerializer(product.tags.all(), many=True)
        
        return Response({
            "message": "บันทึก Tags สำเร็จ",
            "tags": updated_tags.data
        })
        
    except Exception as e:
        return Response({"error": str(e)}, status=500)


# ==========================================
# ⚡ Advanced Tag Logic (Automation & Bulk)
# ==========================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def run_tag_automation_api(request):
    """
    🤖 Smart Tag Automation: Enhanced Rule System
    Rules:
    - Hot Selling (🔥): > 10 sold in 48h
    - Best Seller (🏆): > 50 sold in 30d
    - Last Chance (⌛): Stock < 5
    - New Arrival (🆕): Created in 7d
    - Out of Stock (❌): Stock == 0
    - On Sale (🏷️): price < original_price
    - Flash Sale (⚡): Active in flash sale
    """
    if request.user.role not in ['admin', 'super_admin']:
        return Response({"error": "Unauthorized"}, status=403)
        
    try:
        from .models import Product, Tag, OrderItem, FlashSaleProduct
        from django.db import models
        from django.db.models import Sum, Q, Avg, F
        from django.utils import timezone
        from datetime import timedelta
        
        now = timezone.now()
        stats = {}
        
        # Helper to get/create and ensure Tag meta
        def get_smart_tag(name, color, icon, group="ระบบอัตโนมัติ"):
             tag, _ = Tag.objects.get_or_create(name=name, defaults={'color': color, 'icon': icon, 'group_name': group})
             return tag

        # Define Smart Tags with refined metadata for consistency
        tags_meta = {
            'best_seller': get_smart_tag('Best Seller', '#fbbf24', 'Award', "ยอดนิยม"),
            'hot_selling': get_smart_tag('Hot Selling', '#f87171', 'Zap', "ยอดนิยม"),
            'last_chance': get_smart_tag('Last Chance', '#f472b6', 'Clock', "สถานะสินค้า"),
            'new_arrival': get_smart_tag('New Arrival', '#10b981', 'Sparkles', "สถานะสินค้า"),
            'out_of_stock': get_smart_tag('Out of Stock', '#94a3b8', 'Slash', "สถานะสินค้า"),
            'on_sale': get_smart_tag('On Sale', '#6366f1', 'Percent', "โปรโมชั่น"),
            'flash_sale': get_smart_tag('Flash Sale', '#f59e0b', 'Activity', "โปรโมชั่น"),
        }

        # ---------------------------------------------------------
        # 1. New Arrival (Logic: Created in last 7 days)
        # ---------------------------------------------------------
        seven_days_ago = now - timedelta(days=7)
        new_ids = set(Product.objects.filter(created_at__gte=seven_days_ago).values_list('id', flat=True))
        
        # ---------------------------------------------------------
        # 2. Out of Stock (Logic: stock == 0)
        # ---------------------------------------------------------
        oos_ids = set(Product.objects.filter(stock__lte=0).values_list('id', flat=True))
        
        # ---------------------------------------------------------
        # 3. Last Chance (Logic: 0 < stock < 5)
        # ---------------------------------------------------------
        last_chance_ids = set(Product.objects.filter(stock__gt=0, stock__lt=5).values_list('id', flat=True))

        # ---------------------------------------------------------
        # 4. On Sale (Logic: price < original_price)
        # ---------------------------------------------------------
        on_sale_ids = set(Product.objects.filter(original_price__gt=0).filter(price__lt=F('original_price')).values_list('id', flat=True))

        # ---------------------------------------------------------
        # 5. Flash Sale (Logic: Has active Flash Sale item)
        # ---------------------------------------------------------
        flash_ids = set(FlashSaleProduct.objects.filter(
            flash_sale__is_active=True,
            flash_sale__start_time__lte=now,
            flash_sale__end_time__gte=now
        ).values_list('product_id', flat=True))

        # ---------------------------------------------------------
        # 6. Sales Based (Hot & Best Seller)
        # ---------------------------------------------------------
        # Best Seller (>50 in 30d)
        thirty_days_ago = now - timedelta(days=30)
        best_seller_ids = set(OrderItem.objects.filter(
            order__status__in=['Paid', 'Processing', 'Shipped', 'Completed'],
            order__created_at__gte=thirty_days_ago
        ).values('product_id').annotate(total=Sum('quantity')).filter(total__gte=50).values_list('product_id', flat=True))

        # Hot Selling (>10 in 48h)
        forty_eight_hours_ago = now - timedelta(hours=48)
        hot_selling_ids = set(OrderItem.objects.filter(
            order__status__in=['Paid', 'Processing', 'Shipped', 'Completed'],
            order__created_at__gte=forty_eight_hours_ago
        ).values('product_id').annotate(total=Sum('quantity')).filter(total__gte=10).values_list('product_id', flat=True))

        # ---------------------------------------------------------
        # Final Processing: Bulk Sync Tags
        # ---------------------------------------------------------
        rule_map = {
            'best_seller': best_seller_ids,
            'hot_selling': hot_selling_ids,
            'last_chance': last_chance_ids,
            'new_arrival': new_ids,
            'out_of_stock': oos_ids,
            'on_sale': on_sale_ids,
            'flash_sale': flash_ids
        }

        total_updates = 0
        for key, target_ids in rule_map.items():
            tag = tags_meta[key]
            # 1. Remove from those not in list anymore
            to_remove = Product.objects.filter(tags=tag).exclude(id__in=target_ids)
            for p in to_remove:
                p.tags.remove(tag)
                total_updates += 1
            
            # 2. Add to those in list but not tagged
            to_add = Product.objects.filter(id__in=target_ids).exclude(tags=tag)
            for p in to_add:
                p.tags.add(tag)
                total_updates += 1
            
            stats[tag.name] = len(target_ids)

        return Response({
            "message": "ระบบอัตโนมัติทำงานเสร็จสิ้น",
            "updated_count": total_updates,
            "statistics": stats
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Automation Error: {str(e)}")
        return Response({"error": str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_update_tags_api(request):
    """
    📦 Bulk Management: จัดการ Tag สินค้าทีละหลายรายการ
    """
    if request.user.role not in ['admin', 'super_admin']:
        return Response({"error": "Unauthorized"}, status=403)
        
    try:
        from .models import Product, Tag
        product_ids = request.data.get('product_ids', [])
        tag_id = request.data.get('tag_id')
        action = request.data.get('action', 'add') # 'add' หรือ 'remove'
        
        if not product_ids or not tag_id:
            return Response({"error": "ข้อมูลไม่ครบถ้วน (product_ids, tag_id)"}, status=400)
            
        tag = Tag.objects.get(id=tag_id)
        products = Product.objects.filter(id__in=product_ids)
        
        if action == 'add':
            for p in products:
                p.tags.add(tag)
        else:
            for p in products:
                p.tags.remove(tag)
                
        return Response({
            "message": f"{'เพิ่ม' if action == 'add' else 'ลบ'} Tag '{tag.name}' ให้สินค้า {products.count()} รายการสำเร็จ",
            "action": action
        })
        
    except Tag.DoesNotExist:
        return Response({"error": "ไม่พบ Tag ที่ระบุ"}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def get_products_by_tags_api(request):
    """
    Fetch all products that have ANY of the given tag IDs.
    Returns: List of product data needed for Flash Sale (id, title, thumbnail, price, stock).
    """
    if request.user.role not in ['admin', 'super_admin', 'seller']:
        return Response({"error": "Unauthorized"}, status=403)

    tag_ids = request.data.get('tag_ids', [])
    if not isinstance(tag_ids, list) or not tag_ids:
        return Response({"error": "tag_ids list is required"}, status=400)

    try:
        from .models import Product
        # Filter products that match ANY of the tags
        products = Product.objects.filter(tags__id__in=tag_ids, is_active=True).distinct()
        
        # Serialize only needed fields
        data = []
        for p in products:
            data.append({
                "id": p.id, # Ensure ID is integer
                "title": p.title,
                "price": p.price,
                "stock": p.stock,
                "thumbnail": p.thumbnail.url if p.thumbnail else "/placeholder.png",
                "tags": [{"id": t.id, "name": t.name} for t in p.tags.all()]
            })
            
        return Response(data)

    except Exception as e:
        return Response({"error": str(e)}, status=500)


# ==========================================
# 🎟️ Coupon & Flash Sale APIs
# ==========================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def validate_coupon_api(request):
    import traceback # Add import if needed or assume available
    try:
        code = request.data.get('code')
        try:
            total_amount = float(request.data.get('total_amount', 0))
        except (ValueError, TypeError):
            total_amount = 0.0
        # We might need cart items for product-specific validation (future)
        cart_items = request.data.get('items', []) 

        # 1. Validate via Service
        is_valid, message, coupon = CouponService.validate_coupon(
            user=request.user, 
            coupon_code=code, 
            cart_total=float(total_amount),
            cart_items=cart_items
        )
        
        if not is_valid:
            return Response({"error": message}, status=400)
        
        # 2. Check Stackability (Backend Gatekeeper)
        # If items contain flash sale and coupon is NOT stackable -> Reject?
        # Or frontend handles it? Let's safeguard here.
        # Note: frontend sends total_amount, but maybe not flags.
        # But checking stackability requires knowing if cart has flash sale items.
        # Frontend DOES send 'items', but usually without 'is_flash_sale' flag unless we trust frontend.
        # For now, rely on CouponService calculation or just return valid and let create_order enforce strict rules.
        # Better: If we want to support Stackable in UI, we should return the flag.
            
        # 3. Calculate Discount via Service (Ensures Max Cap logic)
        shipping_cost = 50 # Consistent with create_order
        discount = CouponService.calculate_discount(coupon, total_amount, shipping_cost=shipping_cost)
        
        return Response({
            "valid": True,
            "discount_amount": discount,
            "code": coupon.code,
            "coupon_id": coupon.id, # Added coupon_id
            "type": coupon.discount_type,
            "value": coupon.discount_value,
            "min_spend": coupon.min_spend,
            "max_discount_amount": coupon.max_discount_amount,
            "end_date": coupon.end_date,
            "is_stackable_with_flash_sale": coupon.is_stackable_with_flash_sale # ✅ Send Flag to UI
        })
    except Exception as e:
        print("🔥 CRITICAL ERROR IN VALIDATE COUPON API:")
        print(traceback.format_exc())
        return Response({"error": f"Internal Server Error: {str(e)}"}, status=500)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_public_coupons(request):
    """
    Get all active coupons marked as 'public' for users to choose from.
    """
    try:
        # Simplify: Get ALL active coupons first
        # ✅ Filter out expired coupons (Expired = End Date < Now)
        now = timezone.now()
        coupons = Coupon.objects.filter(active=True, end_date__gte=now)
        
        data = []
        for c in coupons:
            # Basic validation only
            if c.usage_limit > 0 and c.used_count >= c.usage_limit:
                 continue
            
            try:
                data.append({
                    "id": c.id,
                    "code": c.code,
                    "discount_type": c.discount_type,
                    "discount_value": c.discount_value,
                    "min_spend": c.min_spend,
                    "end_date": c.end_date,
                    "start_date": c.start_date, # Added for frontend check
                    "allowed_roles": c.allowed_roles, 
                    "conditions": c.conditions, # ✅ FIXED: Include conditions (e.g. new_user: true)
                    # "description": getattr(c, 'description', None) or ... # Safer
                    "description": f"ส่วนลด {c.discount_value} {'%' if c.discount_type == 'percent' else 'บาท'}"
                })
            except Exception as e:
                print(f"Error processing {c.code}: {e}")
                continue

        return Response(data)
    except Exception as e:
        logger.error(f"Error fetching public coupons: {str(e)}")
        return Response([], status=200)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_active_flash_sales_api(request):
    now = timezone.now()
    active_flash_sales = FlashSale.objects.filter(
        # start_time__lte=now, # ✅ Allow showing upcoming/recent Flash Sales
        end_time__gte=now,
        is_active=True
    ).order_by('-id') # ✅ Show NEWEST created Flash Sale first (instead of soonest ending)
    serializer = FlashSaleSerializer(active_flash_sales, many=True)
    return Response(serializer.data)

# --- Admin Coupon Management ---
@api_view(['GET', 'POST', 'PUT', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def admin_coupon_api(request, coupon_id=None):
    if request.user.role not in ['admin', 'super_admin', 'seller']:
        return Response(status=403)
        
    if request.method == 'GET':
        if coupon_id:
            try:
                c = Coupon.objects.get(id=coupon_id)
                return Response(CouponSerializer(c).data)
            except Coupon.DoesNotExist:
                return Response(status=404)
        else:
            coupons = Coupon.objects.all().order_by('-id')
            return Response(CouponSerializer(coupons, many=True).data)
            
    elif request.method == 'POST':
        # Create or Update
        # If id provided in body, update
        cid = request.data.get('id')
        if cid:
             c = Coupon.objects.get(id=cid)
             s = CouponSerializer(c, data=request.data, partial=True)
        else:
             s = CouponSerializer(data=request.data)
             
        if s.is_valid():
            # ✅ Backend Validation (Enhanced)
            data = s.validated_data
            start = data.get('start_date')
            end = data.get('end_date')
            dtype = data.get('discount_type')
            dval = data.get('discount_value', 0)
            min_spend = data.get('min_spend', 0)
            usage_limit = data.get('usage_limit', 0)

            # 1. Start < End
            if start and end and start > end:
                 return Response({"error": "วันเริ่มต้องมาก่อนวันสิ้นสุด (Start Date must be before End Date)"}, status=400)
            
            # 2. Discount Value > 0 (unless free shipping)
            if dtype != 'free_shipping' and dval <= 0:
                 return Response({"error": "มูลค่าส่วนลดต้องมากกว่า 0 บาท"}, status=400)

            # 3. Percent <= 100
            if dtype == 'percent' and dval > 100:
                 return Response({"error": "ส่วนลดเปอร์เซ็นต์ต้องไม่เกิน 100%"}, status=400)

            # 4. Fixed Discount < Min Spend
            if dtype == 'fixed' and min_spend > 0 and dval >= min_spend:
                 return Response({"error": f"ส่วนลด ({dval}) ต้องน้อยกว่ายอดซื้อขั้นต่ำ ({min_spend})"}, status=400)

            # 5. Usage Limit & Per User > 0
            limit_per_user = data.get('limit_per_user', 1)
            if usage_limit is not None and usage_limit <= 0:
                 return Response({"error": "จำนวนสิทธิ์รวม (Usage Limit) ต้องมากกว่า 0"}, status=400)
            if limit_per_user <= 0:
                 return Response({"error": "จำนวนสิทธิ์ต่อคน (Limit Per User) ต้องมากกว่า 0"}, status=400)

            s.save()
            return Response(s.data)
        return Response(s.errors, status=400)

    elif request.method in ['PUT', 'PATCH']:
        if not coupon_id:
            return Response({"error": "Method requires ID"}, status=400)
        try:
            c = Coupon.objects.get(id=coupon_id)
            # Partial update allowed for PATCH and PUT (for flexibility)
            s = CouponSerializer(c, data=request.data, partial=True)
            if s.is_valid():
                # ✅ Backend Validation (Enhanced for Update)
                data = s.validated_data
                
                # Merge with existing data for missing fields (for robust validation on partial updates)
                start = data.get('start_date', c.start_date)
                end = data.get('end_date', c.end_date)
                dtype = data.get('discount_type', c.discount_type)
                dval = data.get('discount_value', c.discount_value)
                min_spend = data.get('min_spend', c.min_spend)
                usage_limit = data.get('usage_limit', c.usage_limit)
                
                # 1. Start < End
                if start and end and start > end:
                     return Response({"error": "วันเริ่มต้องมาก่อนวันสิ้นสุด"}, status=400)

                # 2. Discount Value > 0
                if dtype != 'free_shipping' and dval <= 0:
                     return Response({"error": "มูลค่าส่วนลดต้องมากกว่า 0 บาท"}, status=400)

                # 3. Percent <= 100
                if dtype == 'percent' and dval > 100:
                     return Response({"error": "ส่วนลดเปอร์เซ็นต์ต้องไม่เกิน 100%"}, status=400)

                # 4. Fixed Discount < Min Spend
                if dtype == 'fixed' and min_spend > 0 and dval >= min_spend:
                     return Response({"error": f"ส่วนลด ({dval}) ต้องน้อยกว่ายอดซื้อขั้นต่ำ ({min_spend})"}, status=400)

                # 5. Usage Limit & Per User > 0
                limit_per_user = data.get('limit_per_user', c.limit_per_user)
                if usage_limit is not None and usage_limit <= 0:
                     return Response({"error": "จำนวนสิทธิ์รวม (Usage Limit) ต้องมากกว่า 0"}, status=400)
                if limit_per_user <= 0:
                     return Response({"error": "จำนวนสิทธิ์ต่อคน (Limit Per User) ต้องมากกว่า 0"}, status=400)

                s.save()
                return Response(s.data)
            return Response(s.errors, status=400)
        except Coupon.DoesNotExist:
            return Response({"error": "Coupon not found"}, status=404)
        
    elif request.method == 'DELETE':
        if not coupon_id: return Response(status=400)
        Coupon.objects.filter(id=coupon_id).delete()
        return Response({"message": "Delete success"})

# --- Admin Flash Sale Management ---
@api_view(['GET', 'POST', 'PUT', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def admin_flash_sale_api(request, fs_id=None):
    # Debug Role
    print(f"DEBUG: Flash Sale Access - User: {request.user.username}, Role: {getattr(request.user, 'role', 'N/A')}, Superuser: {request.user.is_superuser}")
    
    if request.user.role != 'admin' and not request.user.is_superuser and not request.user.is_staff:
        print(f"DEBUG: Access Denied for {request.user.username}")
        return Response(status=403)
        
    if request.method == 'GET':
        if fs_id:
            try:
                fs = FlashSale.objects.get(id=fs_id)
                return Response(FlashSaleSerializer(fs).data)
            except FlashSale.DoesNotExist:
                return Response(status=404)
        else:
            fs = FlashSale.objects.all().order_by('-start_time')
            return Response(FlashSaleSerializer(fs, many=True).data)
            
    elif request.method == 'POST':
        # EXPECT JSON: { "name": "...", "start_time": "...", "end_time": "...", "products": [ {"product_id": 1, "sale_price": 99, "limit": 10}, ... ] }
        data = request.data
        fs_id_param = data.get('id')
        
        # Handle parsed JSON if sent via FormData
        import json
        products_data = data.get('products', [])
        if isinstance(products_data, str):
            try:
                products_data = json.loads(products_data)
            except json.JSONDecodeError:
                products_data = []

        try:
            with transaction.atomic():
                if fs_id_param:
                    try:
                        fs = FlashSale.objects.get(id=fs_id_param)
                        fs.name = data.get('name', fs.name)
                        fs.start_time = data.get('start_time', fs.start_time)
                        fs.end_time = data.get('end_time', fs.end_time)
                        
                        # Fix Boolean
                        raw_active = data.get('is_active', fs.is_active)
                        if isinstance(raw_active, str):
                            raw_active = raw_active.lower() == 'true'
                        fs.is_active = raw_active
                        
                        if 'campaign_id' in data:
                             # ✅ อัปเดต Campaign ที่ Flash Sale นี้สังกัดอยู่
                             # ถ้าส่งเป็น null = ถอดออกจาก Campaign, ถ้าส่งเป็น ID = ย้ายไป Campaign นั้น
                             fs.campaign_id = data.get('campaign_id')
                        
                        fs.save()
                    except FlashSale.DoesNotExist:
                        return Response({"error": "Flash Sale not found"}, status=404)
                    
                    # Clear old products and re-add (Simple Strategy)
                    FlashSaleProduct.objects.filter(flash_sale=fs).delete()
                else:
                    # Fix Boolean
                    raw_active = data.get('is_active', True)
                    if isinstance(raw_active, str):
                        raw_active = raw_active.lower() == 'true'
                        
                    fs = FlashSale.objects.create(
                        name=data['name'],
                        start_time=data['start_time'],
                        end_time=data['end_time'],
                        is_active=raw_active,
                        # ✅ กำหนด Campaign ตั้งแต่ตอนสร้าง (Optional)
                        campaign_id=data.get('campaign_id')
                    )
                
                # Add Products
                # products_data is already parsed above
                print(f"DEBUG: Processing products for create: {products_data}")
                for p_item in products_data:
                    pid = p_item.get('product_id') or p_item.get('id')
                    if not pid:
                        raise ValueError(f"Missing product_id in item: {p_item}")
                        
                    FlashSaleProduct.objects.create(
                        flash_sale=fs,
                        product_id=pid,
                        sale_price=p_item['sale_price'],
                        quantity_limit=p_item.get('quantity_limit', p_item.get('limit', 10)),
                        limit_per_user=p_item.get('limit_per_user', 1)
                    )
                    
            return Response({"message": "Saved successfully"})
        except Exception as e:
            return Response({"error": str(e)}, status=400)
    
    elif request.method in ['PUT', 'PATCH']:
        # RESTful update - fs_id must be provided in URL
        if not fs_id:
            return Response({"error": "Flash Sale ID required in URL"}, status=400)
        
        data = request.data
        
        # Handle parsed JSON if sent via FormData
        import json
        products_data = data.get('products', [])
        if isinstance(products_data, str):
            try:
                products_data = json.loads(products_data)
            except json.JSONDecodeError:
                products_data = []
        
        try:
            with transaction.atomic():
                try:
                    fs = FlashSale.objects.get(id=fs_id)
                    fs.name = data.get('name', fs.name)
                    fs.start_time = data.get('start_time', fs.start_time)
                    fs.end_time = data.get('end_time', fs.end_time)
                    
                    # Fix Boolean
                    raw_active = data.get('is_active', fs.is_active)
                    if isinstance(raw_active, str):
                        raw_active = raw_active.lower() == 'true'
                    fs.is_active = raw_active
                    
                    fs.save()
                except FlashSale.DoesNotExist:
                    return Response({"error": "Flash Sale not found"}, status=404)
                
                # Clear old products and re-add (Simple Strategy)
                FlashSaleProduct.objects.filter(flash_sale=fs).delete()
                
                # Add Products
                # products_data is parsed above
                print(f"DEBUG: Processing products for update: {products_data}")
                for p_item in products_data:
                    pid = p_item.get('product_id') or p_item.get('id')
                    if not pid:
                        raise ValueError(f"Missing product_id in item: {p_item}")

                    FlashSaleProduct.objects.create(
                        flash_sale=fs,
                        product_id=pid,
                        sale_price=p_item['sale_price'],
                        quantity_limit=p_item.get('quantity_limit', p_item.get('limit', 10)),
                        limit_per_user=p_item.get('limit_per_user', 1)
                    )
                    
            return Response({"message": "Updated successfully"})
        except Exception as e:
            return Response({"error": str(e)}, status=400)
            
    elif request.method == 'DELETE':
        if not fs_id:
            return Response({"error": "ID required"}, status=400)
        try:
            FlashSale.objects.filter(id=fs_id).delete()
            return Response({"message": "Deleted successfully"})
        except Exception as e:
            return Response({"error": str(e)}, status=500)

# ==========================================
# 🎯 Flash Sale Campaign API (สำหรับ Timeline & Batch View)
# ==========================================

@api_view(['GET', 'POST', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def admin_campaign_api(request, campaign_id=None):
    """
    🎯 API สำหรับจัดการ Flash Sale Campaign (CRUD)
    
    Campaign คือแคมเปญใหญ่ที่รวม Flash Sale หลายรอบเข้าด้วยกัน
    เช่น "Mega Sale 12.12" อาจมี Flash Sale 5 รอบ (Midnight, Morning, Lunch, Afternoon, Evening)
    
    Methods:
    --------
    GET: ดึงรายการ Campaign ทั้งหมด หรือ 1 Campaign
    POST: สร้าง Campaign ใหม่
    PUT: แก้ไข Campaign ที่มีอยู่
    DELETE: ลบ Campaign
    
    Permissions:
    -----------
    Admin, Super Admin, Seller เท่านั้น
    
    Examples:
    --------
    GET /api/admin/campaigns/           # ดึงทั้งหมด
    GET /api/admin/campaigns/1/         # ดึง Campaign ID 1
    POST /api/admin/campaigns/          # สร้างใหม่
    PUT /api/admin/campaigns/1/         # แก้ไข ID 1
    DELETE /api/admin/campaigns/1/      # ลบ ID 1
    """
    
    # ✅ Check Permission - เฉพาะ Admin/Seller/Super Admin เท่านั้น
    if request.user.role not in ['admin', 'super_admin', 'seller'] and not request.user.is_superuser:
        return Response({"error": "Unauthorized"}, status=403)
    
    # ==========================================
    # 📖 GET - ดึงข้อมูล Campaign
    # ==========================================
    if request.method == 'GET':
        if campaign_id:
            # ดึง 1 Campaign เฉพาะ
            try:
                campaign = FlashSaleCampaign.objects.get(id=campaign_id)
                return Response(FlashSaleCampaignSerializer(campaign).data)
            except FlashSaleCampaign.DoesNotExist:
                return Response({"error": "Campaign not found"}, status=404)
        else:
            # ดึงทั้งหมด (เรียงจากวันที่ล่าสุดและ priority สูงสุดก่อน)
            campaigns = FlashSaleCampaign.objects.all().order_by('-campaign_start', '-priority')
            return Response(FlashSaleCampaignSerializer(campaigns, many=True).data)
    
    # ==========================================
    # ➕ POST - สร้าง Campaign ใหม่
    # ==========================================
    elif request.method == 'POST':
        """
        Expected Payload:
        {
            "name": "Mega Sale 12.12",
            "description": "แคมเปญใหญ่ประจำปี สุดคุ้ม!",
            "campaign_start": "2024-12-12T00:00:00Z",
            "campaign_end": "2024-12-14T23:59:59Z",
            "theme_color": "#ff6600",
            "is_active": true,
            "priority": 10
        }
        """
        try:
            data = request.data
            
            # ✅ Parse Dates to ensure they are datetime objects
            from django.utils.dateparse import parse_datetime
            start_date = parse_datetime(data.get('campaign_start'))
            end_date = parse_datetime(data.get('campaign_end'))

            # 🔥 ตรวจสอบ Campaign ที่ทับซ้อนกัน (Overlap Detection)
            from django.db.models import Q
            overlapping_campaigns = FlashSaleCampaign.objects.filter(
                Q(campaign_start__lte=end_date, campaign_end__gte=start_date)
            )
            
            if overlapping_campaigns.exists():
                # ❌ มี Campaign ทับซ้อน → ห้ามสร้าง
                overlap_list = [{
                    "id": c.id,
                    "name": c.name,
                    "start": c.campaign_start.isoformat(),
                    "end": c.campaign_end.isoformat()
                } for c in overlapping_campaigns]
                
                return Response({
                    "error": "มี Campaign อยู่แล้วในช่วงเวลานี้ (ทับซ้อนไม่ได้)",
                    "overlapping_campaigns": overlap_list
                }, status=400)

            # ✅ สร้าง Campaign Object ใหม่
            campaign = FlashSaleCampaign.objects.create(
                name=data.get('name'),
                description=data.get('description', ''),
                campaign_start=start_date,
                campaign_end=end_date,
                theme_color=data.get('theme_color', '#f97316'),  # สีส้ม (default)
                is_active=data.get('is_active', True),
                priority=data.get('priority', 0)
            )
            
            # ✅ รับรูป Banner (Optional) - ถ้ามีส่งมา
            if 'banner_image' in request.FILES:
                campaign.banner_image = request.FILES['banner_image']
                campaign.save()
            
            # ✅ Return ข้อมูล Campaign ที่สร้างสำเร็จ
            # Create response first
            response_data = FlashSaleCampaignSerializer(campaign).data
            
            # ✅ Handle Flash Sale Assignment (ถ้ามีการส่ง list ของ Flash Sale IDs มา)
            if 'flash_sale_ids' in data and isinstance(data['flash_sale_ids'], list):
                fs_ids = data['flash_sale_ids']
                # 1. Clear existing (optional - depending on logic, here we just add)
                # 2. Add new ones
                FlashSale.objects.filter(id__in=fs_ids).update(campaign=campaign)
                
            return Response(response_data, status=201)
            
        except Exception as e:
            return Response({"error": str(e)}, status=400)
    
    # ==========================================
    # ✏️ PUT - แก้ไข Campaign ที่มีอยู่
    # ==========================================
    elif request.method == 'PUT':
        if not campaign_id:
            return Response({"error": "Campaign ID required in URL"}, status=400)
        
        try:
            campaign = FlashSaleCampaign.objects.get(id=campaign_id)
            data = request.data
            
            # ✅ Update ทุก field (ถ้าไม่ส่งมา                
            # ✅ Parse Dates if provided
            from django.utils.dateparse import parse_datetime
            new_start = campaign.campaign_start
            new_end = campaign.campaign_end
            
            if 'campaign_start' in data:
                new_start = parse_datetime(data['campaign_start'])
            if 'campaign_end' in data:
                new_end = parse_datetime(data['campaign_end'])
            
            # 🔥 ตรวจสอบ Campaign ที่ทับซ้อนกัน (ยกเว้นตัวเอง)
            from django.db.models import Q
            overlapping_campaigns = FlashSaleCampaign.objects.filter(
                Q(campaign_start__lte=new_end, campaign_end__gte=new_start)
            ).exclude(id=campaign_id)  # ยกเว้นตัวเอง
            
            if overlapping_campaigns.exists():
                # ❌ มี Campaign ทับซ้อน → ห้ามแก้ไข
                overlap_list = [{
                    "id": c.id,
                    "name": c.name,
                    "start": c.campaign_start.isoformat(),
                    "end": c.campaign_end.isoformat()
                } for c in overlapping_campaigns]
                
                return Response({
                    "error": "การแก้ไขนี้จะทำให้ทับซ้อนกับ Campaign อื่น (ทับซ้อนไม่ได้)",
                    "overlapping_campaigns": overlap_list
                }, status=400)
            
            # ✅ Apply changes if validation passed
            campaign.campaign_start = new_start
            campaign.campaign_end = new_end
            campaign.name = data.get('name', campaign.name)
            campaign.description = data.get('description', campaign.description)
            campaign.theme_color = data.get('theme_color', campaign.theme_color)
            campaign.is_active = data.get('is_active', campaign.is_active)
            campaign.priority = data.get('priority', campaign.priority)
            
            # ✅ Update รูป Banner (ถ้ามีส่งมาใหม่)
            if 'banner_image' in request.FILES:
                campaign.banner_image = request.FILES['banner_image']
            
            campaign.save()

            # ✅ Handle Flash Sale Assignment (Update)
            if 'flash_sale_ids' in data and isinstance(data['flash_sale_ids'], list):
                fs_ids = data['flash_sale_ids']
                
                # 1. Reset Flash Sales that were in this campaign but NOT in the new list to NULL
                FlashSale.objects.filter(campaign=campaign).exclude(id__in=fs_ids).update(campaign=None)
                
                # 2. Set Campaign for the new list
                FlashSale.objects.filter(id__in=fs_ids).update(campaign=campaign)
            
            return Response(FlashSaleCampaignSerializer(campaign).data)
            
        except FlashSaleCampaign.DoesNotExist:
            return Response({"error": "Campaign not found"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=400)
    
    # ==========================================
    # 🗑️ DELETE - ลบ Campaign
    # ==========================================
    elif request.method == 'DELETE':
        if not campaign_id:
            return Response({"error": "Campaign ID required in URL"}, status=400)
        
        try:
            campaign = FlashSaleCampaign.objects.get(id=campaign_id)
            
            # ✅ ตรวจสอบว่ามี Flash Sale ข้างในหรือไม่
            flash_sale_count = campaign.flash_sales.count()
            
            if flash_sale_count > 0:
                # ⚠️ Warning: ยังมี Flash Sale อยู่ข้างใน
                # แต่ยังลบได้ (Flash Sales จะกลายเป็น orphan - ไม่อยู่ใน Campaign)
                campaign.delete()
                return Response({
                    "message": f"Campaign deleted. {flash_sale_count} Flash Sales are now uncategorized."
                })
            else:
                # ✅ ปลอดภัย ไม่มี Flash Sale ข้างใน ลบได้เลย
                campaign.delete()
                return Response({"message": "Campaign deleted successfully"})
                
        except FlashSaleCampaign.DoesNotExist:
            return Response({"error": "Campaign not found"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_campaign_flash_sales(request, campaign_id):
    """
    📋 ดึง Flash Sales ทั้งหมดที่อยู่ใน Campaign นี้
    
    ใช้สำหรับแสดงรายละเอียด Campaign ว่ามี Flash Sale อะไรบ้างข้างใน
    
    Usage:
    ------
    GET /api/admin/campaigns/1/flash-sales/
    
    Response Example:
    ----------------
    [
        {
            "id": 1,
            "name": "Midnight Sale",
            "start_time": "2024-12-12T00:00:00Z",
            "end_time": "2024-12-12T04:00:00Z",
            "status": "Live",
            "timeline_position_percent": 0,
            "timeline_width_percent": 16.67,
            "timeline_color": "#6366f1",
            ...
        },
        {
            "id": 2,
            "name": "Lunch Flash",
            "start_time": "2024-12-12T11:00:00Z",
            "end_time": "2024-12-12T14:00:00Z",
            ...
        }
    ]
    
    Permissions:
    -----------
    Admin, Super Admin, Seller เท่านั้น
    """
    
    # ✅ Check Permission
    if request.user.role not in ['admin', 'super_admin', 'seller'] and not request.user.is_superuser:
        return Response({"error": "Unauthorized"}, status=403)
    
    try:
        # ดึง Campaign ที่ต้องการ
        campaign = FlashSaleCampaign.objects.get(id=campaign_id)
        
        # ✅ ดึง Flash Sales ทั้งหมดใน Campaign นี้
        # เรียงตาม start_time (แต่ละรอบ) และ display_order
        flash_sales = campaign.flash_sales.all().order_by('start_time', 'display_order')
        
        # Return พร้อม Timeline data (position, width, color)
        return Response(FlashSaleSerializer(flash_sales, many=True).data)
        
    except FlashSaleCampaign.DoesNotExist:
        return Response({"error": "Campaign not found"}, status=404)

# ==========================================
# 🔄 Restored Missing Functions
# ==========================================

@api_view(['POST'])
@permission_classes([AllowAny])
def reset_password_api(request):
    email = request.data.get('email')
    new_password = request.data.get('new_password')
    if not email or not new_password:
        return Response({"error": "Email and new password are required"}, status=400)
    try:
        user = User.objects.get(email=email)
        user.set_password(new_password)
        user.save()
        return Response({"message": "Password updated successfully"})
    except User.DoesNotExist:
        return Response({"error": "User with this email not found"}, status=404)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_notifications(request):
    user = request.user
    notifications = []
    
    # ✅ Filter by Date Since (from Frontend 'Clear' action)
    since_param = request.query_params.get('since')
    since_date = None
    if since_param:
        try:
            from django.utils.dateparse import parse_datetime
            since_date = parse_datetime(since_param)
        except:
            pass
    
    if user.role in ['admin', 'super_admin', 'seller']:
        # Admin Notifications: New Orders, Low Stock
        orders_query = Order.objects.filter(status='Pending')
        if since_date:
            orders_query = orders_query.filter(created_at__gt=since_date)
            
        recent_orders = orders_query.order_by('-created_at')[:5]
        for o in recent_orders:
            notifications.append({
                "id": f"order_{o.id}",
                "title": "คำสั่งซื้อใหม่",
                "message": f"คุณได้รับคำสั่งซื้อใหม่จาก {o.customer_name} ยอด {o.total_price} บาท",
                "time": o.created_at.strftime("%d/%m %H:%M"),
                "timestamp": o.created_at.isoformat(), 
                "type": "order"
            })
        
        # Low Stock - Only show if NOT filtering by recency (to prevent persistent alerts after clear)
        if not since_date:
            low_stock = Product.objects.filter(stock__lte=5, is_active=True)[:5]
            for p in low_stock:
                notifications.append({
                    "id": f"stock_{p.id}",
                    "title": "สินค้าใกล้หมด",
                    "message": f"สินค้า '{p.title}' เหลือเพียง {p.stock} ชิ้น",
                    "time": timezone.now().strftime("%d/%m %H:%M"),
                    "timestamp": timezone.now().isoformat(),
                    "type": "alert"
                })
    else:
        # Customer Notifications: Order Status Changes
        orders_query = Order.objects.filter(user=user)
        if since_date:
            orders_query = orders_query.filter(updated_at__gt=since_date)
            
        my_recent_orders = orders_query.order_by('-updated_at')[:10]
        for o in my_recent_orders:
            if o.status != 'Pending':
                notifications.append({
                    "id": f"cust_order_{o.id}",
                    "title": f"อัปเดตคำสั่งซื้อ #{o.id}",
                    "message": f"คำสั่งซื้อของคุณเปลี่ยนสถานะเป็น: {o.status}",
                    "time": o.updated_at.strftime("%d/%m %H:%M"),
                    "timestamp": o.updated_at.isoformat(), 
                    "type": "success" if o.status == 'Completed' else "info"
                })
                
    return Response(notifications)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_admin_logs(request):
    if request.user.role not in ['admin', 'super_admin', 'seller']:
        return Response(status=403)
    logs = AdminLog.objects.all().order_by('-timestamp')[:100]
    data = [{
        "id": l.id,
        "admin": l.admin.username if l.admin else "System",
        "action": l.action,
        "details": l.details,
        "ip_address": l.ip_address,
        "timestamp": l.timestamp.strftime("%Y-%m-%d %H:%M:%S")
    } for l in logs]
    return Response(data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_all_users(request):
    if request.user.role != 'admin':
        return Response(status=403)
    users = User.objects.all().order_by('-id')
    data = [{
        "id": u.id,
        "username": u.username,
        "email": u.email,
        "first_name": u.first_name,
        "last_name": u.last_name,
        "phone": u.phone,
        "address": u.address,
        "role": u.role,  # Frontend expects 'role' (or we align frontend to this)
        "is_active": u.is_active,
        "avatar": u.image.url if u.image else "",
        "date_joined": u.date_joined.strftime("%Y-%m-%d")
    } for u in users]
    return Response(data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_product_api(request):
    if request.user.role not in ['admin', 'super_admin', 'seller']:
        return Response(status=403)
    try:
        data = request.data
        
        # ✅ Resolve Category FIRST (to avoid null error on create)
        category_obj = None
        if 'category' in data and data['category']:
            category_obj, _ = Category.objects.get_or_create(name=data['category'])
            
        product = Product.objects.create(
            title=data.get('title'),
            description=data.get('description', ''),
            price=data.get('price'),
            original_price=data.get('original_price'),
            stock=data.get('stock', 0),
            brand=data.get('brand', ''),
            sku=data.get('sku'),
            category=category_obj, # ✅ Pass immediately
            is_active=True
        )
            
        if 'thumbnail' in request.FILES:
            product.thumbnail = request.FILES['thumbnail']
            product.save()

        # ✅ NEW: Handle multiple gallery images (รูปภาพเพิ่มเติม)
        # รับไฟล์หลายรูปจาก key 'images'
        images = request.FILES.getlist('images')
        for img in images:
            ProductImage.objects.create(product=product, image_url=img)
            
        return Response({"message": "Product added with images", "id": product.id}, status=201)
    except Exception as e:
        return Response({"error": str(e)}, status=400)

@api_view(['PUT', 'POST'])
@permission_classes([IsAuthenticated])
def edit_product_api(request, product_id):
    if request.user.role not in ['admin', 'super_admin', 'seller']:
        return Response(status=403)
    try:
        product = Product.objects.get(id=product_id)
        data = request.data
        
        if 'title' in data: product.title = data['title']
        if 'description' in data: product.description = data['description']
        if 'price' in data and data['price'] != '':
            try:
                product.price = float(data['price'])
            except ValueError:
                pass # Keep old price if invalid

        if 'stock' in data and data['stock'] != '':
             try:
                product.stock = int(data['stock'])
             except ValueError:
                pass

        if 'brand' in data: product.brand = data['brand']
        
        if 'category' in data and data['category']:
            cat, _ = Category.objects.get_or_create(name=data['category'])
            product.category = cat
            
        if 'thumbnail' in request.FILES:
            product.thumbnail = request.FILES['thumbnail']

        # ✅ NEW: Add new gallery images during edit (เพิ่มรูปใหม่เข้าไปใน Gallery)
        # Support both 'images' and 'new_gallery_images' keys
        gallery_files = request.FILES.getlist('images') + request.FILES.getlist('new_gallery_images')
        
        for img in gallery_files:
            ProductImage.objects.create(product=product, image_url=img)
            
        product.save()
        return Response({"message": "Product updated"})
    except Product.DoesNotExist:
        return Response(status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=400)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_product_api(request, product_id):
    if request.user.role not in ['admin', 'super_admin', 'seller']:
        return Response(status=403)
    try:
        product = Product.objects.get(id=product_id)
        product.is_active = False # Soft delete
        product.save()
        return Response({"message": "Product deactivated"})
    except Product.DoesNotExist:
        return Response(status=404)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_product_image_api(request, image_id):
    if request.user.role not in ['admin', 'super_admin', 'seller']:
        return Response(status=403)
    try:
        ProductImage.objects.filter(id=image_id).delete()
        return Response({"message": "Image deleted"})
    except Exception as e:
        return Response({"error": str(e)}, status=400)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_categories(request):
    categories = Category.objects.all().order_by('name')
    data = [{"id": c.id, "name": c.name} for c in categories]
    return Response(data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_all_stock_history(request):
    try:
        if not hasattr(request.user, 'role') or request.user.role not in ['admin', 'super_admin', 'seller']:
            return Response({'error': 'Unauthorized'}, status=403)
        
        # 1. รับค่า params
        search_query = request.query_params.get('search', '').strip()
        action_filter = request.query_params.get('action', 'all')
        category_filter = request.query_params.get('category', 'all')

        # 2. Query Base
        queryset = StockHistory.objects.select_related('product', 'created_by').all().order_by('-created_at')

        # 3. Apply Filters
        if search_query:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(product__title__icontains=search_query) |
                Q(note__icontains=search_query)
            )
        
        if action_filter != 'all':
            queryset = queryset.filter(action=action_filter)
            
        if category_filter != 'all':
            # Assuming product has category field (ForeignKey or CharField)
            # If relation: product__category__name or product__category
            # Check model first usually, but assuming generic 'category' field on Product or relationship.
            # Safe bet: Filter by product__category__name if it's a relation, or product__category if CharField.
            # checking common pattern in this project...
            queryset = queryset.filter(product__category__name=category_filter)

        # Limit results
        history = queryset[:100]

        data = []
        for h in history:
            try:
                prod_title = h.product.title if h.product else "Unknown Product"
            except:
                prod_title = "Error Loading Product"

            data.append({
                "id": h.id,
                "product": prod_title,
                "change": h.change_quantity,
                "remaining": h.remaining_stock,
                "action": h.action,
                "note": h.note or "",
                "user": h.created_by.username if h.created_by else "System",
                "date": h.created_at.strftime("%Y-%m-%d %H:%M")
            })

        return Response(data)

    except Exception as e:
        print(f"❌ Error fetching stock history: {str(e)}")
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_stock_history(request, product_id):
    if request.user.role not in ['admin', 'super_admin', 'seller']:
        return Response(status=403)
    try:
        product = Product.objects.get(id=product_id)
        history = StockHistory.objects.filter(product=product).order_by('-created_at')
        data = [{
            "id": h.id,
            "product": h.product.title,
            "change": h.change_quantity,
            "remaining": h.remaining_stock,
            "action": h.action,
            "timestamp": h.created_at.strftime("%Y-%m-%d %H:%M")
        } for h in history]
        return Response(data)
    except Product.DoesNotExist:
        return Response({"error": "Product not found"}, status=404)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_related_products(request, product_id):
    try:
        product = Product.objects.get(id=product_id)
        related = Product.objects.filter(category=product.category, is_active=True).exclude(id=product_id)[:4]
        data = [{
            "id": p.id,
            "title": p.title,
            "price": p.price,
            "thumbnail": p.thumbnail.url if p.thumbnail else "",
            "tags": [{"id": t.id, "name": t.name, "icon": t.icon, "color": t.color, "slug": t.slug} for t in p.tags.all()]
        } for p in related]
        return Response(data)
    except:
        return Response([])

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def submit_review(request):
    try:
        user = request.user
        product_id = request.data.get('product_id')
        rating = request.data.get('rating')
        comment = request.data.get('comment', '')
        image = request.FILES.get('image', None)  # ✅ รับรูปภาพจาก FormData
        
        product = Product.objects.get(id=product_id)
        Review.objects.create(
            user=user, 
            product=product, 
            rating=rating, 
            comment=comment,
            image=image  # ✅ บันทึกรูปภาพ
        )
        
        # Update product average rating
        all_reviews = product.reviews.all()
        avg_rating = sum([r.rating for r in all_reviews]) / all_reviews.count()
        product.rating = avg_rating
        product.save()
        
        return Response({"message": "Review submitted"})
    except Exception as e:
        return Response({"error": str(e)}, status=400)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_order_api(request, order_id):
    if request.user.role not in ['admin', 'super_admin', 'seller']:
        return Response(status=403)
    Order.objects.filter(id=order_id).delete()
    return Response({"message": "Order deleted"})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_promptpay_qr_api(request):
    try:
        amount = float(request.data.get('amount', 0))
        if amount <= 0:
             return Response({"error": "Invalid amount"}, status=400)
             
        # Generate Fake PromptPay Payload
        phone = "0812345678" 
        payload = f"00020101021129370016A00000067701011101130066{phone[1:]}5802TH5303764540{int(amount * 100):04d}0000" 
        
        return Response({"payload": payload})
    except Exception as e:
        return Response({"error": str(e)}, status=400)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_orders_api_v4(request):
    try:
        # 1. Role Check
        if not hasattr(request.user, 'role') or request.user.role not in ['seller', 'admin', 'super_admin']: 
            return Response({"error": "Unauthorized Access"}, status=403)
        
        # 2. Base Query
        orders = Order.objects.all().order_by('-created_at')
        
        # 3. Filters
        search = request.query_params.get('search', '')
        if search:
            orders = orders.filter(
                Q(id__icontains=search) | 
                Q(customer_name__icontains=search) |
                Q(customer_tel__icontains=search)
            )

        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        if start_date and end_date:
            # Using string formatting for simple date filter if range fails
            orders = orders.filter(created_at__date__range=[start_date, end_date])
        
        status = request.query_params.get('status', 'ทั้งหมด')
        if status != 'ทั้งหมด':
            orders = orders.filter(status=status)
            
        province = request.query_params.get('province', 'ทั้งหมด')
        if province != 'ทั้งหมด':
            orders = orders.filter(shipping_province=province)

        # 4. Serialization
        data = []
        for o in orders:
            try:
                # items = [f"{i.product.title} (x{i.quantity})" for i in o.items.all()] # OLD
                items = []
                for i in o.items.all():
                    if i.product:
                        items.append({
                            "product": i.product.title,
                            "quantity": i.quantity,
                            "price": i.price_at_purchase
                        })
                    else:
                        items.append({
                            "product": "สินค้าถูกลบ (Deleted)",
                            "quantity": i.quantity,
                            "price": i.price_at_purchase if hasattr(i, 'price_at_purchase') else 0
                        })

                data.append({
                    "id": o.id,
                    "date": o.created_at.strftime("%d/%m/%Y %H:%M"),
                    "customer": o.customer_name,
                    "tel": o.customer_tel,
                    "province": o.shipping_province,
                    "items": items,
                    "total_price": o.total_price,
                    "status": o.status,
                    "slip_image": o.slip_image.url if o.slip_image else None,
                    "transfer_amount": o.transfer_amount,
                    "transfer_date": o.transfer_date.strftime("%d/%m/%Y %H:%M") if o.transfer_date else None,
                    "bank_name": o.bank_name,
                    "transfer_account_number": o.transfer_account_number
                })
            except Exception as e:
                print(f"Error serializing order {o.id}: {e}")
                continue
            
        print(f"DEBUG: Admin Orders V3 returning {len(data)} items")
        return Response(data)
        
    except Exception as e:
        print(f"CRITICAL ERROR in admin_orders_api_v3: {e}")
        return Response({"error": str(e)}, status=500)
    
    return Response({"error": "Unexpected Fallthrough"}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_admin_logs(request):
    if request.user.role not in ['admin', 'super_admin']:
        return Response(status=403)
        
    logs = AdminLog.objects.all().order_by('-timestamp')[:100]
    data = []
    
    for l in logs:
        # Simple Category inference
        category = 'Other'
        action_lower = l.action.lower()
        if 'login' in action_lower or 'logout' in action_lower: category = 'Auth'
        elif 'order' in action_lower: category = 'Order'
        elif 'product' in action_lower: category = 'Product'
        elif 'user' in action_lower: category = 'User'
        
        data.append({
            "id": l.id,
            "admin": l.admin.username if l.admin else "System",
            "action": l.action,
            "date": l.timestamp.strftime("%d/%m/%Y %H:%M"),
            "category": category
        })
        
    return Response(data)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_all_products_simple(request):
    # Public Access allowed (Dropdown uses public info)
    products = Product.objects.filter(is_active=True).order_by('title')
    data = [{
        "id": p.id,
        "title": p.title,
        "original_price": p.price,
        "thumbnail": p.thumbnail.url if p.thumbnail else ""
    } for p in products]
    return Response(data)

# ==========================================
# 🎫 Coupon System API (Collection)
# ==========================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def collect_coupon_api(request, coupon_id):
    """
    เก็บคูปอง (Collect Coupon)
    """
    try:
        # ✅ 0. Role Restriction (ป้องกัน Admin/Seller แย่งคูปองลูกค้า)
        if request.user.role in ['admin', 'super_admin', 'seller']:
             return Response({'message': 'Admin และ Seller ไม่สามารถเก็บคูปองได้ครับ (สำหรับลูกค้าเท่านั้น)'}, status=status.HTTP_403_FORBIDDEN)

        coupon = Coupon.objects.get(pk=coupon_id)
        
        # 1. เช็คว่าเคยเก็บไปแล้วหรือยัง (ตาม limit_per_user)
        collected_count = UserCoupon.objects.filter(user=request.user, coupon=coupon).count()
        if collected_count >= coupon.limit_per_user:
             return Response({'message': 'คุณเก็บคูปองนี้ครบจำนวนสิทธิ์แล้ว'}, status=status.HTTP_400_BAD_REQUEST)
            
        # 2. เช็คว่าคูปองหมดอายุหรือยัง
        now = timezone.now()
        if not coupon.active or not (coupon.start_date <= now <= coupon.end_date):
             return Response({'message': 'คูปองหมดอายุหรือยังไม่เปิดให้เก็บ'}, status=status.HTTP_400_BAD_REQUEST)
        
        # 3. เช็คสิทธิ์รวม (Total Supply)
        # เช็คจำนวนคนที่เก็บไปแล้ว
        total_collected = UserCoupon.objects.filter(coupon=coupon).count()
        limit = max(coupon.total_supply, coupon.usage_limit)
        if total_collected >= limit:
             return Response({'message': 'คูปองหมดแล้ว!'}, status=status.HTTP_400_BAD_REQUEST)

        # 4. สร้าง UserCoupon
        UserCoupon.objects.create(
            user=request.user, 
            coupon=coupon,
            status='active'
        )
        
        return Response({'message': 'เก็บคูปองสำเร็จ!'}, status=status.HTTP_200_OK)

    except Coupon.DoesNotExist:
        return Response({'message': 'ไม่พบคูปอง'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Collect Coupon Error: {str(e)}")
        return Response({'message': 'เกิดข้อผิดพลาดในการเก็บคูปอง'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_my_coupons_api(request):
    """
    ดึงคูปองของฉัน (My Coupons)
    """
    try:
        # ดึงคูปองที่เก็บไว้ เรียงตามวันที่เก็บล่าสุด
        user_coupons = UserCoupon.objects.filter(user=request.user).select_related('coupon').order_by('-collected_at')
        
        data = []
        for uc in user_coupons:
            c = uc.coupon
            
            # คำนวณวันหมดอายุจริง (Expiry Date)
            expiry_date = c.end_date
            
            data.append({
                "id": c.id,
                "user_coupon_id": uc.id,
                "code": c.code,
                "name": c.name,
                "description": c.description,
                "discount_type": c.discount_type,
                "discount_value": c.discount_value,
                "min_spend": c.min_spend,
                "expiry_date": expiry_date,
                "is_expired": expiry_date < timezone.now(),
                "status": uc.status, # active, used, expired
                "is_used": uc.status == 'used' or uc.used_at is not None
            })
            
        return Response(data)
    except Exception as e:
        logger.error(f"Get My Coupons Error: {str(e)}")
        return Response({'message': 'เกิดข้อผิดพลาดในการดึงข้อมูลคูปอง'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
